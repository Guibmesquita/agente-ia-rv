"""
Endpoints para Campanhas Ativas e Templates de Mensagem.
Permite criar campanhas de disparo em massa para assessores.
"""
import json
import io
import re
import asyncio
import random
from datetime import datetime
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel, validator
from database.database import get_db, SessionLocal
from database.models import MessageTemplate, Campaign, CampaignDispatch, CampaignStatus, Assessor, CampaignStructure
from api.endpoints.auth import require_role
from database.models import User

DISPATCH_DELAY_MIN = 3.0
DISPATCH_DELAY_MAX = 10.0
MAX_RETRY_ATTEMPTS = 3
RETRY_DELAY_SECONDS = 3.0

cancelled_campaigns: dict = {}


def get_random_dispatch_delay() -> float:
    """Retorna um delay aleatório entre 3 e 10 segundos para simular envio manual."""
    return random.uniform(DISPATCH_DELAY_MIN, DISPATCH_DELAY_MAX)


def normalize_code(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)
    if isinstance(value, int):
        return str(value)
    s = str(value).strip()
    if not s:
        return ""
    try:
        num = float(s)
        if num == int(num):
            return str(int(num))
    except (ValueError, TypeError):
        pass
    return s


def format_cell_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.startswith("R$") or not stripped:
            return value
        if stripped.startswith("0") and not stripped.startswith("0."):
            return value
        try:
            num = float(stripped)
        except (ValueError, TypeError):
            return value
        if num == int(num):
            return str(int(num))
        return f"{num:.2f}"
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return f"{value:.2f}"
    if isinstance(value, int):
        return str(value)
    return str(value)

router = APIRouter(prefix="/api/campaigns", tags=["campaigns"])

# Task #221 — Router auxiliar para expor o estado do motor sob o prefixo
# `/api/cadence` (contrato exigido pela task: GET /api/cadence/engine-state)
# em paralelo ao endpoint canônico em `/api/campaigns/cadence/engine-state`.
cadence_router = APIRouter(prefix="/api/cadence", tags=["cadence"])


def translate_error_to_natural_language(error_code: str, error_msg: str, phone: str = "") -> str:
    """
    Traduz códigos de erro técnicos para linguagem natural amigável.
    """
    translations = {
        "TIMEOUT": f"O servidor do WhatsApp demorou muito para responder. Isso pode indicar que o serviço está sobrecarregado ou indisponível.",
        "CONNECTION_ERROR": "Não foi possível conectar ao servidor do WhatsApp. Verifique se as credenciais Z-API estão corretas e o serviço está online.",
        "HTTP_401": "Credenciais inválidas. O Token ou Client-Token do Z-API pode estar incorreto.",
        "HTTP_403": "Acesso negado. Verifique as permissões da sua instância Z-API.",
        "HTTP_404": "Endpoint não encontrado. Verifique se o Instance ID está correto.",
        "HTTP_500": "Erro interno no servidor do WhatsApp. O serviço Z-API pode estar com problemas.",
        "HTTP_502": "O servidor do WhatsApp está temporariamente indisponível (Bad Gateway).",
        "HTTP_503": "O servidor do WhatsApp está em manutenção ou sobrecarregado.",
        "API_ERROR": f"A API do WhatsApp retornou um erro: {error_msg}",
        "HTTP_ERROR": f"Erro de comunicação com o servidor: {error_msg}",
    }
    
    if error_code in translations:
        base_msg = translations[error_code]
    elif error_code.startswith("HTTP_"):
        base_msg = f"O servidor retornou código de erro {error_code.replace('HTTP_', '')}: {error_msg}"
    else:
        base_msg = f"Erro ao enviar mensagem: {error_msg}"
    
    if "not registered" in error_msg.lower() or "number not exist" in error_msg.lower():
        base_msg = f"O número {phone} não está registrado no WhatsApp ou está inativo."
    elif "session not found" in error_msg.lower():
        base_msg = "A sessão do WhatsApp não foi encontrada. É necessário reconectar o WhatsApp no painel Z-API."
    elif "not connected" in error_msg.lower():
        base_msg = "O WhatsApp não está conectado. Verifique se o celular está online e conectado à internet."
    elif "invalid phone" in error_msg.lower() or "invalid number" in error_msg.lower():
        base_msg = f"O número {phone} está em formato inválido. Verifique se está no padrão correto (ex: 5511999999999)."
    
    return base_msg


def template_has_required_variables(template: str) -> bool:
    """
    Verifica se o template contém as variáveis obrigatórias.
    Aceita variações com e sem espaços, e com uma ou duas chaves.
    """
    if not template:
        return False
    
    # Padrões aceitos para nome_assessor
    has_nome = any([
        "{{nome_assessor}}" in template,
        "{{ nome_assessor }}" in template,
        "{nome_assessor}" in template,
    ])
    
    # Padrões aceitos para lista_clientes
    has_lista = any([
        "{{lista_clientes}}" in template,
        "{{ lista_clientes }}" in template,
        "{lista_clientes}" in template,
    ])
    
    return has_nome and has_lista


# Mensagem padrao usada quando nenhum template e selecionado
DEFAULT_TEMPLATE_CONTENT = """Ola, {{nome_assessor}}!

Seguem as recomendacoes de troca de ativos para seus clientes:

{{lista_clientes}}

Por favor, entre em contato com cada cliente para alinhar as operacoes.

Atenciosamente,
Equipe de Gestao"""


class TemplateCreate(BaseModel):
    name: str
    content: str
    description: Optional[str] = None
    attachment_url: Optional[str] = None
    attachment_type: Optional[str] = None
    attachment_filename: Optional[str] = None
    variables_used: Optional[List[str]] = None


class TemplateUpdate(BaseModel):
    name: Optional[str] = None
    content: Optional[str] = None
    description: Optional[str] = None
    is_active: Optional[bool] = None
    attachment_url: Optional[str] = None
    attachment_type: Optional[str] = None
    attachment_filename: Optional[str] = None
    variables_used: Optional[List[str]] = None


class ColumnMapping(BaseModel):
    assessor_id: str
    assessor_email: str
    client_id: str
    ativo_saida: str
    valor_saida: str
    ativo_compra: str
    valor_compra: str


class CustomFieldMapping(BaseModel):
    column_name: str
    variable_name: str


class CampaignCreate(BaseModel):
    name: str
    template_id: Optional[int] = None


class CampaignUpdate(BaseModel):
    name: Optional[str] = None
    template_id: Optional[int] = None
    column_mapping: Optional[dict] = None
    custom_fields_mapping: Optional[dict] = None


class CampaignDispatchRequest(BaseModel):
    campaign_id: int


def require_admin_or_gestao():
    """Dependency that requires admin or gestao_rv role."""
    return require_role(["admin", "gestao_rv"])


@router.get("/templates")
async def list_templates(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    """Lista todos os templates de mensagem ativos."""
    templates = db.query(MessageTemplate).filter(
        MessageTemplate.is_active == 1
    ).order_by(MessageTemplate.name).all()
    
    return [
        {
            "id": t.id,
            "name": t.name,
            "content": t.content,
            "description": t.description,
            "attachment_url": t.attachment_url,
            "attachment_type": t.attachment_type,
            "attachment_filename": t.attachment_filename,
            "variables_used": json.loads(t.variables_used) if t.variables_used else [],
            "created_at": t.created_at.isoformat() if t.created_at else None,
            "updated_at": t.updated_at.isoformat() if t.updated_at else None
        }
        for t in templates
    ]


@router.get("/templates/{template_id}")
async def get_template(
    template_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    """Busca um template por ID."""
    template = db.query(MessageTemplate).filter(MessageTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template não encontrado")
    
    return {
        "id": template.id,
        "name": template.name,
        "content": template.content,
        "description": template.description,
        "is_active": template.is_active == 1,
        "attachment_url": template.attachment_url,
        "attachment_type": template.attachment_type,
        "attachment_filename": template.attachment_filename,
        "variables_used": json.loads(template.variables_used) if template.variables_used else [],
        "created_at": template.created_at.isoformat() if template.created_at else None
    }


def extract_variables_from_content(content: str) -> List[str]:
    """Extrai variáveis no formato {{variavel}} do conteúdo."""
    pattern = r'\{\{\s*([a-zA-Z_][a-zA-Z0-9_]*)\s*\}\}'
    matches = re.findall(pattern, content)
    return list(set(matches))


@router.post("/templates")
async def create_template(
    data: TemplateCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    """Cria um novo template de mensagem."""
    variables = data.variables_used or extract_variables_from_content(data.content)
    
    template = MessageTemplate(
        name=data.name,
        content=data.content,
        description=data.description,
        attachment_url=data.attachment_url,
        attachment_type=data.attachment_type,
        attachment_filename=data.attachment_filename,
        variables_used=json.dumps(variables),
        created_by=int(current_user.id),
        is_active=1
    )
    db.add(template)
    db.commit()
    db.refresh(template)
    
    return {
        "id": template.id,
        "name": template.name,
        "variables_used": variables,
        "message": "Template criado com sucesso"
    }


@router.put("/templates/{template_id}")
async def update_template(
    template_id: int,
    data: TemplateUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    """Atualiza um template existente."""
    template = db.query(MessageTemplate).filter(MessageTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template não encontrado")
    
    if data.name is not None:
        template.name = data.name
    if data.content is not None:
        template.content = data.content
        template.variables_used = json.dumps(extract_variables_from_content(data.content))
    if data.description is not None:
        template.description = data.description
    if data.is_active is not None:
        template.is_active = 1 if data.is_active else 0
    if data.attachment_url is not None:
        template.attachment_url = data.attachment_url
    if data.attachment_type is not None:
        template.attachment_type = data.attachment_type
    if data.attachment_filename is not None:
        template.attachment_filename = data.attachment_filename
    if data.variables_used is not None:
        template.variables_used = json.dumps(data.variables_used)
    
    db.commit()
    db.refresh(template)
    
    return {
        "message": "Template atualizado com sucesso",
        "variables_used": json.loads(template.variables_used) if template.variables_used else []
    }


@router.delete("/templates/{template_id}")
async def delete_template(
    template_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    """Remove um template (soft delete)."""
    template = db.query(MessageTemplate).filter(MessageTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template não encontrado")
    
    template.is_active = 0
    db.commit()
    
    return {"message": "Template removido com sucesso"}


# ─────────────────────────────────────────────────────────────────────────────
# Task #226 — Campanha de teste multi-canal
# ─────────────────────────────────────────────────────────────────────────────

class TestSendRequest(BaseModel):
    phones: List[str]
    # Task #256: múltiplos canais — cada canal envia o template para todos os números.
    channel_ids: List[int]
    # Task #259: mensagem livre composta no wizard (substitui template_id + preview_name)
    message_text: str


def _normalize_phone_for_test(raw: str) -> tuple:
    """Normaliza número e retorna (normalized_str, is_valid_bool)."""
    clean = re.sub(r"\D", "", raw.strip())
    if len(clean) in (10, 11):
        clean = "55" + clean
    valid = len(clean) in (12, 13) and clean.isdigit()
    return clean, valid


def _render_preview_content(content: str, preview_name: str) -> str:
    """Substitui variáveis conhecidas por valores de preview; desconhecidas ficam [VAR]."""
    known = {
        "nome_assessor": preview_name,
        "nome": preview_name,
        "codigo_ai": "TST001",
        "lista_clientes": "[Lista de clientes — disparo de teste]",
        "unidade": "Unidade Teste",
        "equipe": "Equipe Teste",
        "broker": "Broker Teste",
        "data": datetime.now().strftime("%d/%m/%Y"),
    }
    for key, val in known.items():
        content = re.sub(r"\{\{\s*" + re.escape(key) + r"\s*\}\}", val, content, flags=re.IGNORECASE)
    content = re.sub(r"\{\{\s*([a-zA-Z_][a-zA-Z0-9_]*)\s*\}\}", r"[\1]", content)
    return content


@router.get("/template-preview")
async def preview_template_render(
    template_id: int,
    preview_name: str = "Assessor Teste",
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao()),
):
    """
    Task #226 — Renderiza o conteúdo de um template substituindo variáveis por
    valores de preview fictícios. Variáveis desconhecidas ficam como [VARIAVEL].
    """
    template = db.query(MessageTemplate).filter(
        MessageTemplate.id == template_id,
        MessageTemplate.is_active == 1,
    ).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template não encontrado")

    rendered = _render_preview_content(str(template.content), preview_name)
    variables = json.loads(template.variables_used) if template.variables_used else []

    return {
        "template_id": template.id,
        "template_name": template.name,
        "rendered_content": rendered,
        "attachment_url": template.attachment_url,
        "attachment_type": template.attachment_type,
        "attachment_filename": template.attachment_filename,
        "variables_used": variables,
    }


@router.post("/test-send")
async def test_send_stream(
    data: TestSendRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao()),
):
    """
    Task #226 — Disparo de teste multi-canal via SSE (POST + fetch ReadableStream).
    Envia mensagens de texto para números livres usando os canais Z-API selecionados.
    Nenhum registro é criado em campaigns ou campaign_dispatches.
    Delay aleatório 3-8s entre envios por canal (anti-bloqueio).
    Logar cada envio com tag [TEST-SEND].

    Task #256 — Suporte a múltiplos canais: channel_ids aceita uma lista de IDs.
    O template é enviado de cada canal para todos os números de destino (canal × número).
    Ordem de envio: canal 1 → todos os números → canal 2 → todos os números → …
    Eventos SSE incluem channel_id, channel_label e channel_index para o frontend
    distinguir por canal.
    """
    import logging as _logging
    _log = _logging.getLogger(__name__)

    # Validações iniciais (todas antes de iniciar o stream)
    if len(data.phones) > 20:
        raise HTTPException(status_code=400, detail="Limite de 20 números por disparo de teste")

    if not data.channel_ids:
        raise HTTPException(status_code=400, detail="Selecione ao menos um canal")

    # Deduplica mantendo a ordem de seleção (evita envios duplicados via mesmo canal)
    seen_ids: set = set()
    unique_channel_ids = []
    for cid in data.channel_ids:
        if cid not in seen_ids:
            seen_ids.add(cid)
            unique_channel_ids.append(cid)

    # Task #259: valida mensagem livre
    message_text = data.message_text.strip()
    if not message_text:
        raise HTTPException(status_code=400, detail="A mensagem não pode estar vazia")
    if len(message_text) > 4096:
        raise HTTPException(status_code=400, detail="A mensagem excede o limite de 4096 caracteres")

    from database.models import ZAPIChannel as _ZCh
    from services.whatsapp_client import ZAPIClient as _ZC

    # Valida cada canal e pré-instancia clientes (antes do stream para falhar rápido)
    channels_ready = []
    for cid in unique_channel_ids:
        ch = db.query(_ZCh).filter(_ZCh.id == cid).first()
        if not ch:
            raise HTTPException(status_code=404, detail=f"Canal #{cid} não encontrado")
        if not ch.is_active:
            raise HTTPException(status_code=400, detail=f"Canal '{ch.label or ch.name}' está inativo")
        client = _ZC(instance_id=ch.instance_id, token=ch.token, client_token=ch.client_token)
        if not client.is_configured():
            raise HTTPException(
                status_code=400,
                detail=f"Canal '{ch.label or ch.name}' não está configurado (credenciais incompletas)",
            )
        label = ch.label or ch.name or f"Canal #{cid}"
        channels_ready.append((ch.id, label, client))

    # Normalizar e filtrar números válidos
    normalized = []
    for raw in data.phones:
        norm, valid = _normalize_phone_for_test(raw)
        if valid and norm not in normalized:
            normalized.append(norm)

    if not normalized:
        raise HTTPException(status_code=400, detail="Nenhum número válido encontrado")

    # Mensagem já pronta — sem substituição de variáveis (texto livre)
    rendered_content = message_text

    total = len(normalized) * len(channels_ready)

    async def _stream():
        sent_total = 0
        failed_total = 0
        global_idx = 0
        # Resultado por canal para o evento final
        by_channel = []

        for ch_idx, (ch_id, ch_label, client) in enumerate(channels_ready):
            sent_ch = 0
            failed_ch = 0
            for ph_idx, phone in enumerate(normalized):
                global_idx += 1
                try:
                    result = await client.send_text(phone, rendered_content)
                    success = result.get("success", False)
                    error = None if success else result.get("error", "Erro desconhecido")
                    if success:
                        sent_ch += 1
                        sent_total += 1
                        _log.info(f"[TEST-SEND] status=sent canal={ch_label} phone={phone} ({global_idx}/{total})")
                        # Task #261 — persistir mensagem de teste na conversa do destinatário.
                        # Usa SessionLocal dedicado para não conflitar com a sessão da request SSE.
                        _db2 = SessionLocal()
                        try:
                            from api.endpoints.whatsapp_webhook import save_message_zapi as _swz
                            import uuid as _uuid
                            _swz(
                                _db2,
                                message_id=f"TEST-{_uuid.uuid4().hex[:16]}",
                                zaap_id=None,
                                phone=phone,
                                direction="outbound",
                                message_type="text",
                                from_me=True,
                                body=rendered_content,
                                sender_type="bot",
                                channel_id=ch_id,
                                is_test_dispatch=True,
                            )
                        except Exception as _pe:
                            _log.warning(f"[TEST-SEND] Falha ao persistir na conversa de {phone}: {_pe}")
                        finally:
                            _db2.close()
                    else:
                        failed_ch += 1
                        failed_total += 1
                        _log.info(f"[TEST-SEND] status=failed canal={ch_label} phone={phone} error={error!r} ({global_idx}/{total})")
                except Exception as exc:
                    success = False
                    error = str(exc)
                    failed_ch += 1
                    failed_total += 1
                    _log.info(f"[TEST-SEND] status=exception canal={ch_label} phone={phone} exc={error!r} ({global_idx}/{total})")

                progress_evt = json.dumps({
                    "type": "progress",
                    "channel_id": ch_id,
                    "channel_label": ch_label,
                    "channel_index": ch_idx,
                    "phone_index": ph_idx,
                    "phone": phone,
                    "status": "sent" if success else "failed",
                    "error": error,
                    "index": global_idx,
                    "total": total,
                })
                yield f"data: {progress_evt}\n\n"

                # Delay anti-bloqueio entre envios do mesmo canal (exceto após o último)
                if ph_idx < len(normalized) - 1:
                    await asyncio.sleep(random.uniform(3.0, 8.0))

            by_channel.append({
                "channel_id": ch_id,
                "channel_label": ch_label,
                "sent": sent_ch,
                "failed": failed_ch,
            })
            # Pausa entre canais (menor — são instâncias separadas)
            if ch_idx < len(channels_ready) - 1:
                await asyncio.sleep(random.uniform(1.0, 3.0))

        done_evt = json.dumps({
            "type": "done",
            "sent": sent_total,
            "failed": failed_total,
            "total": total,
            "by_channel": by_channel,
        })
        yield f"data: {done_evt}\n\n"

    return StreamingResponse(
        _stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ─────────────────────────────────────────────────────────────────────────────

import os
import uuid
from pathlib import Path

ATTACHMENTS_DIR = Path("uploads/attachments")
ALLOWED_ATTACHMENT_TYPES = {
    'image/jpeg': 'image',
    'image/png': 'image',
    'image/gif': 'image',
    'image/webp': 'image',
    'video/mp4': 'video',
    'video/quicktime': 'video',
    'audio/mpeg': 'audio',
    'audio/mp3': 'audio',
    'audio/ogg': 'audio',
    'audio/wav': 'audio',
    'application/pdf': 'document',
    'application/msword': 'document',
    'application/vnd.openxmlformats-officedocument.wordprocessingml.document': 'document',
    'application/vnd.ms-excel': 'document',
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': 'document',
    'text/plain': 'document',
}


@router.post("/attachments/upload")
async def upload_attachment(
    file: UploadFile = File(...),
    current_user: User = Depends(require_admin_or_gestao())
):
    """
    Faz upload de anexo (imagem, vídeo, áudio ou documento) para campanhas/templates.
    Retorna a URL do arquivo para ser usada no template ou campanha.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="Nome do arquivo inválido")
    
    content_type = file.content_type or 'application/octet-stream'
    
    if content_type not in ALLOWED_ATTACHMENT_TYPES:
        allowed_types = list(ALLOWED_ATTACHMENT_TYPES.keys())
        raise HTTPException(
            status_code=400, 
            detail=f"Tipo de arquivo não permitido. Tipos aceitos: imagem (JPEG, PNG, GIF, WebP), vídeo (MP4), áudio (MP3, OGG, WAV), documento (PDF, DOC, DOCX, XLS, XLSX, TXT)"
        )
    
    attachment_type = ALLOWED_ATTACHMENT_TYPES[content_type]
    
    file_ext = Path(file.filename).suffix.lower()
    unique_filename = f"{uuid.uuid4().hex}{file_ext}"
    
    ATTACHMENTS_DIR.mkdir(parents=True, exist_ok=True)
    file_path = ATTACHMENTS_DIR / unique_filename
    
    try:
        contents = await file.read()
        max_size = 50 * 1024 * 1024
        if len(contents) > max_size:
            raise HTTPException(status_code=400, detail="Arquivo muito grande. Tamanho máximo: 50MB")
        
        with open(file_path, 'wb') as f:
            f.write(contents)
        
        file_url = f"/uploads/attachments/{unique_filename}"
        
        return {
            "success": True,
            "url": file_url,
            "type": attachment_type,
            "filename": file.filename,
            "size": len(contents),
            "message": "Arquivo enviado com sucesso"
        }
    except HTTPException:
        raise
    except Exception as e:
        if file_path.exists():
            file_path.unlink()
        raise HTTPException(status_code=500, detail=f"Erro ao salvar arquivo: {str(e)}")


@router.delete("/attachments")
async def delete_attachment(
    url: str,
    current_user: User = Depends(require_admin_or_gestao())
):
    """Remove um anexo pelo URL."""
    if not url or not url.startswith("/uploads/attachments/"):
        raise HTTPException(status_code=400, detail="URL de anexo inválida")
    
    filename = url.replace("/uploads/attachments/", "")
    file_path = ATTACHMENTS_DIR / filename
    
    if file_path.exists():
        file_path.unlink()
        return {"message": "Anexo removido com sucesso"}
    else:
        raise HTTPException(status_code=404, detail="Anexo não encontrado")


@router.post("/upload")
async def upload_campaign_file(
    file: UploadFile = File(...),
    campaign_name: str = Form(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    """
    Faz upload de arquivo CSV/Excel e retorna as colunas para mapeamento.
    Cria uma campanha em rascunho.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="Nome do arquivo inválido")
    
    filename = file.filename.lower()
    if not (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
        raise HTTPException(status_code=400, detail="Formato inválido. Use CSV ou Excel.")
    
    try:
        contents = await file.read()
        
        if filename.endswith('.csv'):
            import csv
            text = contents.decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(text))
            columns = reader.fieldnames or []
            rows = list(reader)
        else:
            import pandas as pd
            import openpyxl
            import math
            
            pct_columns = set()
            try:
                wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
                ws = wb.active
                if ws:
                    col_names = []
                    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False), None)
                    if header_row:
                        col_names = [cell.value for cell in header_row]
                    for row_cells in ws.iter_rows(min_row=2, max_row=min(11, ws.max_row or 2), values_only=False):
                        for cell in row_cells:
                            fmt = cell.number_format or ""
                            if "%" in fmt and cell.value is not None:
                                col_idx = cell.column - 1
                                if col_idx < len(col_names) and col_names[col_idx]:
                                    pct_columns.add(str(col_names[col_idx]))
                wb.close()
            except Exception as e:
                print(f"[UPLOAD] Warning: could not detect percentage columns: {e}")
            
            if pct_columns:
                print(f"[UPLOAD] Detected percentage columns: {pct_columns}")
            
            df = pd.read_excel(io.BytesIO(contents))
            columns = df.columns.tolist()
            rows = df.to_dict('records')
            
            sanitized = []
            for row in rows:
                clean = {}
                for k, v in row.items():
                    try:
                        if pd.isna(v):
                            clean[k] = None
                            continue
                    except (ValueError, TypeError):
                        pass
                    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                        clean[k] = None
                    elif str(k) in pct_columns and isinstance(v, (int, float)):
                        pct_val = v * 100
                        if pct_val == int(pct_val):
                            clean[k] = f"{int(pct_val)}%"
                        else:
                            clean[k] = f"{pct_val:.2f}%"
                    else:
                        clean[k] = v
                sanitized.append(clean)
            rows = sanitized
        
        campaign = Campaign(
            name=campaign_name,
            status=CampaignStatus.DRAFT.value,
            original_filename=file.filename,
            total_recommendations=len(rows),
            processed_data=json.dumps(rows, default=str),
            created_by=int(current_user.id)
        )
        db.add(campaign)
        db.commit()
        db.refresh(campaign)
        
        suggested_mapping = suggest_column_mapping(columns)
        
        return {
            "campaign_id": campaign.id,
            "filename": file.filename,
            "columns": columns,
            "row_count": len(rows),
            "suggested_mapping": suggested_mapping,
            "preview": rows[:5] if rows else []
        }
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao processar arquivo: {str(e)}")


def suggest_column_mapping(columns: List[str]) -> dict:
    """Sugere mapeamento automático baseado nos nomes das colunas."""
    mapping = {}
    columns_lower = {c.lower().strip(): c for c in columns}
    
    field_patterns = {
        "assessor_id": ["assessor", "id_assessor", "cod_assessor", "codigo_assessor", "advisor"],
        "assessor_email": ["email", "email_assessor", "e-mail", "mail", "assessor_email"],
        "client_id": ["cliente", "id_cliente", "cod_cliente", "codigo_cliente", "client"],
        "ativo_saida": ["ativo_saida", "saida", "venda", "papel_saida", "ticker_saida"],
        "valor_saida": ["valor_saida", "vl_saida", "valor_venda"],
        "ativo_compra": ["ativo_compra", "compra", "papel_compra", "ticker_compra"],
        "valor_compra": ["valor_compra", "vl_compra", "valor_entrada"]
    }
    
    for field, patterns in field_patterns.items():
        for pattern in patterns:
            if pattern in columns_lower:
                mapping[field] = columns_lower[pattern]
                break
    
    return mapping


class CampaignFromBaseRequest(BaseModel):
    name: str
    assessor_ids: List[int]


@router.post("/from-base")
async def create_campaign_from_base(
    data: CampaignFromBaseRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    """
    Cria uma campanha a partir de assessores selecionados da base.
    Não requer mapeamento de colunas pois os dados já estão estruturados.
    """
    if not data.assessor_ids:
        raise HTTPException(status_code=400, detail="Nenhum assessor selecionado")
    
    assessores = db.query(Assessor).filter(Assessor.id.in_(data.assessor_ids)).all()
    
    if len(assessores) != len(data.assessor_ids):
        raise HTTPException(status_code=400, detail="Um ou mais assessores não foram encontrados")
    
    assessor_data = []
    for a in assessores:
        assessor_data.append({
            "id": a.id,
            "codigo_ai": a.codigo_ai,
            "nome": a.nome,
            "email": a.email,
            "telefone_whatsapp": a.telefone_whatsapp,
            "unidade": a.unidade,
            "equipe": a.equipe,
            "broker_responsavel": a.broker_responsavel
        })
    
    campaign = Campaign(
        name=data.name,
        status=CampaignStatus.DRAFT.value,
        original_filename=None,
        total_recommendations=0,
        total_assessors=len(assessores),
        source_type="base",
        processed_data=json.dumps(assessor_data, default=str),
        created_by=int(current_user.id)
    )
    db.add(campaign)
    db.commit()
    db.refresh(campaign)
    
    return {
        "campaign_id": campaign.id,
        "assessor_count": len(assessores),
        "message": "Campanha criada com sucesso"
    }


class CampaignMappingRequest(BaseModel):
    column_mapping: Optional[dict] = None
    custom_fields_mapping: Optional[dict] = None
    message_template: Optional[str] = None
    message_blocks: Optional[dict] = None
    message_header: Optional[str] = None
    message_content_template: Optional[str] = None
    message_footer: Optional[str] = None
    group_by_client: Optional[bool] = False
    content_line_template: Optional[str] = None
    assessor_code_column: Optional[str] = None
    attachment_url: Optional[str] = None
    attachment_type: Optional[str] = None
    attachment_filename: Optional[str] = None


@router.put("/{campaign_id}/mapping")
async def update_campaign_mapping(
    campaign_id: int,
    request: CampaignMappingRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    """Atualiza o mapeamento de colunas de uma campanha."""
    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campanha não encontrada")
    
    if request.column_mapping:
        campaign.column_mapping = json.dumps(request.column_mapping)
    
    if request.custom_fields_mapping:
        campaign.custom_fields_mapping = json.dumps(request.custom_fields_mapping)
    
    if request.message_template:
        campaign.message_content = request.message_template
    
    if request.message_blocks:
        campaign.message_header = request.message_blocks.get("header", "")
        campaign.message_content_template = request.message_blocks.get("content", "")
        campaign.message_footer = request.message_blocks.get("footer", "")
    
    if request.message_header is not None:
        campaign.message_header = request.message_header
    if request.message_content_template is not None:
        campaign.message_content_template = request.message_content_template
    if request.message_footer is not None:
        campaign.message_footer = request.message_footer
    
    if request.content_line_template:
        campaign.message_content_template = request.content_line_template
    
    campaign.group_by_client = 1 if request.group_by_client else 0
    
    if request.assessor_code_column:
        mapping = json.loads(campaign.column_mapping or "{}")
        mapping["codigo_ai"] = request.assessor_code_column
        campaign.column_mapping = json.dumps(mapping)
    
    if request.attachment_url is not None:
        campaign.attachment_url = request.attachment_url
    if request.attachment_type is not None:
        campaign.attachment_type = request.attachment_type
    if request.attachment_filename is not None:
        campaign.attachment_filename = request.attachment_filename
    
    db.commit()
    
    return {"message": "Mapeamento atualizado com sucesso"}


@router.put("/{campaign_id}/template")
async def set_campaign_template(
    campaign_id: int,
    template_id: int = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    """Define o template de mensagem para a campanha."""
    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campanha não encontrada")
    
    if template_id:
        template = db.query(MessageTemplate).filter(MessageTemplate.id == template_id).first()
        if not template:
            raise HTTPException(status_code=404, detail="Template não encontrado")
        campaign.template_id = template_id
    
    db.commit()
    
    return {"message": "Template definido com sucesso"}


class MessageBlocksModel(BaseModel):
    header: Optional[str] = ""
    content: Optional[str] = ""
    footer: Optional[str] = ""


class CustomTemplateRequest(BaseModel):
    content: Optional[str] = None
    message_blocks: Optional[MessageBlocksModel] = None
    message_header: Optional[str] = None
    message_content_template: Optional[str] = None
    message_footer: Optional[str] = None
    content_line_template: Optional[str] = None
    group_by_client: Optional[bool] = False
    client_id_column: Optional[str] = None
    attachment_url: Optional[str] = None
    attachment_type: Optional[str] = None
    attachment_filename: Optional[str] = None


@router.put("/{campaign_id}/custom-template")
async def set_custom_template(
    campaign_id: int,
    request: CustomTemplateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    """Define um template customizado (editado) para a campanha."""
    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campanha não encontrada")
    
    header = None
    content_template = None
    footer = None
    
    if request.message_blocks:
        header = request.message_blocks.header or ""
        content_template = request.message_blocks.content or request.content_line_template or ""
        footer = request.message_blocks.footer or ""
    elif request.message_header is not None or request.message_content_template is not None or request.message_footer is not None:
        header = request.message_header or ""
        content_template = request.message_content_template or request.content_line_template or ""
        footer = request.message_footer or ""
    
    if header is not None or content_template is not None or footer is not None:
        campaign.message_header = header
        campaign.message_content_template = content_template or request.content_line_template
        campaign.message_footer = footer
        campaign.group_by_client = 1 if request.group_by_client else 0
        campaign.client_id_column = request.client_id_column
        
        full_content_parts = []
        if header:
            full_content_parts.append(header)
        if content_template:
            full_content_parts.append(content_template)
        if footer:
            full_content_parts.append(footer)
        campaign.custom_template_content = "\n\n".join(full_content_parts) if full_content_parts else request.content
    elif request.content:
        campaign.custom_template_content = request.content
    
    if request.attachment_url is not None:
        campaign.attachment_url = request.attachment_url
    if request.attachment_type is not None:
        campaign.attachment_type = request.attachment_type
    if request.attachment_filename is not None:
        campaign.attachment_filename = request.attachment_filename
    
    db.commit()
    
    return {"message": "Template customizado salvo com sucesso"}


@router.get("/{campaign_id}/attachment-check")
async def check_campaign_attachment(
    campaign_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    """
    Verifica se o arquivo de anexo de uma campanha está acessível antes do disparo.
    Retorna se o arquivo existe no filesystem local ou é uma URL pública válida.
    """
    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campanha não encontrada")

    attachment_url = campaign.attachment_url
    attachment_filename = campaign.attachment_filename or ""

    if not attachment_url:
        return {
            "has_attachment": False,
            "accessible": True,
            "filename": None,
            "message": "Campanha sem anexo"
        }

    url = attachment_url.strip()

    if url.startswith(("http://", "https://", "data:")):
        return {
            "has_attachment": True,
            "accessible": True,
            "filename": attachment_filename,
            "message": "Anexo disponível via URL pública"
        }

    if not url.startswith("/"):
        url = "/" + url

    _UPLOADS_ROOT = os.path.realpath(
        os.path.join(os.getcwd(), "uploads", "attachments")
    )
    raw_local = url.lstrip("/")
    candidate = os.path.realpath(os.path.join(os.getcwd(), raw_local))

    if not candidate.startswith(_UPLOADS_ROOT + os.sep) and candidate != _UPLOADS_ROOT:
        return {
            "has_attachment": True,
            "accessible": False,
            "filename": attachment_filename,
            "message": "Caminho do arquivo fora do diretório permitido"
        }

    if os.path.isfile(candidate):
        return {
            "has_attachment": True,
            "accessible": True,
            "filename": attachment_filename,
            "message": "Anexo disponível no servidor"
        }

    return {
        "has_attachment": True,
        "accessible": False,
        "filename": attachment_filename,
        "message": f"Arquivo \"{attachment_filename or raw_local}\" não encontrado no servidor. Faça o upload novamente antes de disparar."
    }


@router.get("/{campaign_id}/preview")
async def preview_campaign(
    campaign_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    """
    Gera preview das mensagens agrupadas por assessor.
    Aplica a lógica de agrupamento e substitui variáveis.
    """
    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campanha não encontrada")
    
    source_type = getattr(campaign, 'source_type', 'upload') or 'upload'
    
    if source_type == "base":
        return await preview_campaign_from_base(campaign, db)
    
    # Usa template customizado se existir, senao template salvo, senao mensagem padrao
    template_content = DEFAULT_TEMPLATE_CONTENT
    template_name = "Mensagem Padrao"
    
    if campaign.custom_template_content:
        candidate = str(campaign.custom_template_content)
        if template_has_required_variables(candidate):
            template_content = candidate
            template_name = "Mensagem Editada"
        else:
            print(f"[PREVIEW] Template customizado não contém variáveis obrigatórias, usando padrão")
    elif campaign.template_id:
        template = db.query(MessageTemplate).filter(MessageTemplate.id == campaign.template_id).first()
        if template:
            candidate = str(template.content)
            if template_has_required_variables(candidate):
                template_content = candidate
                template_name = str(template.name)
            else:
                print(f"[PREVIEW] Template salvo não contém variáveis obrigatórias, usando padrão")
    
    try:
        column_mapping = json.loads(str(campaign.column_mapping)) if campaign.column_mapping else {}
        custom_mapping = json.loads(str(campaign.custom_fields_mapping)) if campaign.custom_fields_mapping else {}
        data = json.loads(str(campaign.processed_data)) if campaign.processed_data else []
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Erro nos dados da campanha")
    
    print(f"[DEBUG PREVIEW] campaign_id={campaign_id}")
    print(f"[DEBUG PREVIEW] column_mapping={column_mapping}")
    print(f"[DEBUG PREVIEW] custom_mapping={custom_mapping}")
    print(f"[DEBUG PREVIEW] data rows count={len(data)}")
    if data:
        print(f"[DEBUG PREVIEW] first row keys={list(data[0].keys())}")
        print(f"[DEBUG PREVIEW] first row sample={data[0]}")
    
    if not column_mapping:
        raise HTTPException(status_code=400, detail="Mapeamento de colunas não definido")
    
    grouped = group_recommendations_by_assessor(data, column_mapping, custom_mapping, db)
    print(f"[DEBUG PREVIEW] grouped keys={list(grouped.keys())}")
    if grouped:
        first_key = list(grouped.keys())[0]
        print(f"[DEBUG PREVIEW] first assessor data={grouped[first_key]}")
    
    campaign.total_assessors = len(grouped)
    db.commit()
    
    content_line_template = campaign.message_content_template or ""
    
    messages = []
    first_example = None
    
    is_grouped = bool(campaign.group_by_client)
    
    for idx, (assessor_id, assessor_data) in enumerate(grouped.items()):
        message = build_message(template_content, assessor_data, custom_mapping, content_line_template, group_by_client=is_grouped)
        messages.append({
            "assessor_id": assessor_id,
            "assessor_name": assessor_data.get("nome_assessor", ""),
            "assessor_phone": assessor_data.get("telefone", ""),
            "client_count": len(assessor_data.get("clients", {})),
            "recommendation_count": assessor_data.get("total_recommendations", 0),
            "message_preview": message
        })
        
        if idx == 0:
            clients_data = assessor_data.get("clients", {})
            
            header_template = campaign.message_header or ""
            content_template = campaign.message_content_template or ""
            footer_template = campaign.message_footer or ""
            
            import unicodedata
            
            def normalize_var_name(name: str) -> str:
                """Remove acentos e normaliza nome de variável."""
                if not name:
                    return ""
                normalized = unicodedata.normalize('NFKD', str(name))
                ascii_text = normalized.encode('ASCII', 'ignore').decode('ASCII')
                return ascii_text.lower().replace(' ', '_').replace('-', '_')
            
            def is_formatted_currency(value) -> bool:
                """Verifica se o valor já está formatado como moeda."""
                if not isinstance(value, str):
                    return False
                return value.strip().startswith("R$")
            
            def replace_vars(text, data):
                result = text
                normalized_data = {}
                for key, value in data.items():
                    if isinstance(value, (dict, list)):
                        continue
                    norm_key = normalize_var_name(key)
                    str_value = format_cell_value(value)
                    
                    if norm_key in normalized_data:
                        existing = normalized_data[norm_key]
                        if is_formatted_currency(existing) and not is_formatted_currency(str_value):
                            continue
                    
                    normalized_data[norm_key] = str_value
                    normalized_data[str(key)] = str_value
                
                import re
                pattern = r'\{\{\s*([^}]+?)\s*\}\}'
                def replacer(match):
                    var_name = match.group(1).strip()
                    normalized_name = normalize_var_name(var_name)
                    if normalized_name in normalized_data:
                        return normalized_data[normalized_name]
                    if var_name in normalized_data:
                        return normalized_data[var_name]
                    return match.group(0)
                
                result = re.sub(pattern, replacer, result)
                return result
            
            nome_assessor = assessor_data.get("nome_assessor", "") or assessor_data.get("nome", "")
            primeiro_nome = str(nome_assessor).split()[0] if nome_assessor else ""
            
            extra_vars = {
                "nome_assessor": nome_assessor,
                "assessor": nome_assessor,
                "primeiro_nome": primeiro_nome,
                "nome": nome_assessor
            }
            
            header_rendered = replace_vars(header_template, assessor_data)
            header_rendered = replace_vars(header_rendered, extra_vars)
            
            content_lines = []
            total_recs = 0
            is_grouped = bool(campaign.group_by_client)
            for client_id, client_info in clients_data.items():
                recommendations = client_info.get("recommendations", [])
                if recommendations:
                    if is_grouped:
                        content_lines.append(f"Cliente {client_id}")
                    for rec in recommendations:
                        line = replace_vars(content_template, rec)
                        line = replace_vars(line, assessor_data)
                        line = replace_vars(line, extra_vars)
                        if line.strip():
                            if is_grouped:
                                content_lines.append(f"- {line}")
                            else:
                                content_lines.append(line)
                        total_recs += 1
                    if is_grouped:
                        content_lines.append("")
            
            footer_rendered = replace_vars(footer_template, assessor_data)
            footer_rendered = replace_vars(footer_rendered, extra_vars)
            
            first_example = {
                "assessor_name": assessor_data.get("nome_assessor", "Assessor"),
                "assessor_id": assessor_id,
                "recommendation_count": total_recs,
                "header_rendered": header_rendered,
                "content_lines": content_lines,
                "footer_rendered": footer_rendered
            }
    
    return {
        "campaign_id": campaign.id,
        "total_assessors": len(messages),
        "total_recommendations": campaign.total_recommendations,
        "messages": messages[:5],
        "template_name": template_name,
        "first_example": first_example
    }


async def preview_campaign_from_base(campaign, db: Session):
    """
    Gera preview para campanhas baseadas em assessores selecionados da base.
    Não tem recomendações de ativos, apenas lista de assessores para disparo.
    Mostra exatamente o que o usuário preencheu, sem dados de exemplo.
    """
    header_template = campaign.message_header or ""
    content_template = campaign.message_content_template or ""
    footer_template = campaign.message_footer or ""
    
    template_name = "Mensagem Personalizada"
    if not header_template and not content_template and not footer_template:
        if campaign.custom_template_content:
            content_template = str(campaign.custom_template_content)
            template_name = "Mensagem Editada"
        elif campaign.template_id:
            template = db.query(MessageTemplate).filter(MessageTemplate.id == campaign.template_id).first()
            if template:
                content_template = str(template.content)
                template_name = str(template.name)
    
    try:
        data = json.loads(str(campaign.processed_data)) if campaign.processed_data else []
    except json.JSONDecodeError:
        data = []
    
    first_example = {
        "assessor_name": "Assessor Exemplo",
        "header_rendered": header_template.strip() if header_template.strip() else None,
        "content_lines": [content_template.strip()] if content_template.strip() else [],
        "footer_rendered": footer_template.strip() if footer_template.strip() else None,
        "header_empty": not header_template.strip(),
        "content_empty": not content_template.strip(),
        "footer_empty": not footer_template.strip()
    }
    
    messages = []
    
    for idx, assessor in enumerate(data):
        variables = build_assessor_variables(assessor)
        
        header_rendered = replace_variables_generic(header_template, variables)
        content_rendered = replace_variables_generic(content_template, variables)
        footer_rendered = replace_variables_generic(footer_template, variables)
        
        message_parts = []
        if header_rendered.strip():
            message_parts.append(header_rendered.strip())
        if content_rendered.strip():
            message_parts.append(content_rendered.strip())
        if footer_rendered.strip():
            message_parts.append(footer_rendered.strip())
        
        full_message = "\n\n".join(message_parts) if message_parts else "(Mensagem vazia)"
        
        if idx == 0:
            first_example = {
                "assessor_name": assessor.get("nome", "Assessor"),
                "header_rendered": header_rendered.strip() if header_rendered.strip() else None,
                "content_lines": [content_rendered.strip()] if content_rendered.strip() else [],
                "footer_rendered": footer_rendered.strip() if footer_rendered.strip() else None,
                "header_empty": not header_template.strip(),
                "content_empty": not content_template.strip(),
                "footer_empty": not footer_template.strip()
            }
        
        messages.append({
            "assessor_id": str(assessor.get("id", "")),
            "assessor_name": assessor.get("nome", ""),
            "assessor_phone": assessor.get("telefone_whatsapp", ""),
            "client_count": 0,
            "recommendation_count": 0,
            "message_preview": full_message
        })
    
    return {
        "campaign_id": campaign.id,
        "total_assessors": len(messages),
        "total_recommendations": 0,
        "messages": messages[:5],
        "template_name": template_name,
        "source_type": "base",
        "first_example": first_example
    }


def group_recommendations_by_assessor(data: List[dict], mapping: dict, custom_mapping: dict, db: Session) -> dict:
    """
    Agrupa recomendações por assessor.
    
    Suporta dois modos:
    1. Modo legado: usa assessor_id/assessor_email para identificar assessores e constrói lista_clientes
    2. Modo codigo_ai: usa codigo_ai para vincular com base interna e disponibiliza variáveis da planilha
    
    Retorna um dicionário com dados do assessor e recomendações agrupadas.
    """
    grouped = {}
    
    col_codigo_ai = mapping.get("codigo_ai", "")
    col_assessor = mapping.get("assessor_id", "")
    col_assessor_email = mapping.get("assessor_email", "")
    col_client = mapping.get("client_id", "")
    col_ativo_saida = mapping.get("ativo_saida", "")
    col_valor_saida = mapping.get("valor_saida", "")
    col_ativo_compra = mapping.get("ativo_compra", "")
    col_valor_compra = mapping.get("valor_compra", "")
    
    def find_column(row, possible_names):
        """Busca uma coluna pelos possíveis nomes (com e sem acentos)."""
        for name in possible_names:
            if name in row:
                return name
        return None
    
    ativo_saida_names = ["Ativo saída", "Ativo Saída", "ativo_saida", "ativo saida", "ATIVO SAÍDA", "Ativo Saida"]
    valor_saida_names = ["Valor saída", "Valor Saída", "valor_saida", "valor saida", "VALOR SAÍDA", "Valor Saida"]
    ativo_compra_names = ["Ativo compra", "Ativo Compra", "ativo_compra", "ativo compra", "ATIVO COMPRA", "Ativo Entrada"]
    valor_compra_names = ["Valor compra", "Valor Compra", "valor_compra", "valor compra", "VALOR COMPRA", "Valor Entrada"]
    
    use_codigo_ai_mode = bool(col_codigo_ai) and not col_assessor
    
    print(f"[GROUPING] Column mapping: codigo_ai={col_codigo_ai}, assessor={col_assessor}")
    print(f"[GROUPING] Mode: {'codigo_ai' if use_codigo_ai_mode else 'legacy'}")
    print(f"[GROUPING] Total rows to process: {len(data)}")
    
    for idx, row in enumerate(data):
        if use_codigo_ai_mode:
            codigo_ai_val = row.get(col_codigo_ai, "")
            if codigo_ai_val is None:
                codigo_ai_val = ""
            key = normalize_code(codigo_ai_val)
        else:
            assessor_val = row.get(col_assessor, "")
            if assessor_val is None:
                assessor_val = ""
            key = normalize_code(assessor_val)
        
        if not key:
            print(f"[GROUPING] Row {idx}: No key found, skipping")
            continue
        
        if key not in grouped:
            if use_codigo_ai_mode:
                assessor = db.query(Assessor).filter(Assessor.codigo_ai == key).first()
                if not assessor:
                    stripped_key = re.sub(r'^[A-Za-z]+', '', key)
                    if stripped_key and stripped_key != key:
                        assessor = db.query(Assessor).filter(Assessor.codigo_ai == stripped_key).first()
                        if assessor:
                            print(f"[GROUPING] Found assessor by stripping prefix: {key} -> {stripped_key}")
            else:
                assessor = None
                email_from_sheet = ""
                if col_assessor_email:
                    email_val = row.get(col_assessor_email, "")
                    if email_val:
                        email_from_sheet = str(email_val).strip()
                        assessor = db.query(Assessor).filter(Assessor.email == email_from_sheet).first()
                
                if not assessor:
                    try:
                        assessor_id_int = int(key)
                        assessor = db.query(Assessor).filter(Assessor.id == assessor_id_int).first()
                    except (ValueError, TypeError):
                        pass
                
                if not assessor:
                    assessor = db.query(Assessor).filter(
                        (Assessor.telefone_whatsapp == key) |
                        (Assessor.nome.ilike(f"%{key}%"))
                    ).first()
            
            if assessor:
                grouped[key] = {
                    "codigo_ai": assessor.codigo_ai or "",
                    "nome": assessor.nome or "",
                    "email": assessor.email or "",
                    "telefone_whatsapp": assessor.telefone_whatsapp or "",
                    "telefone": assessor.telefone_whatsapp or "",
                    "unidade": assessor.unidade or "",
                    "equipe": assessor.equipe or "",
                    "broker_responsavel": assessor.broker_responsavel or "",
                    "nome_assessor": assessor.nome or "",
                    "assessor_id": key,
                    "email_assessor": assessor.email or "",
                    "clients": {},
                    "total_recommendations": 0,
                    "custom_fields": {},
                    "spreadsheet_data": {}
                }
            else:
                grouped[key] = {
                    "codigo_ai": key if use_codigo_ai_mode else "",
                    "nome": key if not use_codigo_ai_mode else "",
                    "email": "",
                    "telefone_whatsapp": "",
                    "telefone": "",
                    "unidade": "",
                    "equipe": "",
                    "broker_responsavel": "",
                    "nome_assessor": key if not use_codigo_ai_mode else "",
                    "assessor_id": key,
                    "email_assessor": "",
                    "clients": {},
                    "total_recommendations": 0,
                    "custom_fields": {},
                    "spreadsheet_data": {}
                }
                print(f"[GROUPING] Warning: Assessor not found for key={key}")
            
            print(f"[GROUPING] New assessor: {key} -> nome={grouped[key]['nome']}")
        
        for col_name, col_val in row.items():
            if col_name not in grouped[key]["spreadsheet_data"]:
                grouped[key]["spreadsheet_data"][col_name] = col_val
        
        if isinstance(custom_mapping, dict):
            for col_name, var_name in custom_mapping.items():
                if col_name in row and row[col_name]:
                    grouped[key]["custom_fields"][var_name] = str(row[col_name])
        elif isinstance(custom_mapping, list):
            for custom_item in custom_mapping:
                if isinstance(custom_item, dict):
                    col_name = custom_item.get("column_name", "")
                    var_name = custom_item.get("variable_name", "")
                    if col_name in row and row[col_name]:
                        grouped[key]["custom_fields"][var_name] = str(row[col_name])
        
        if col_client:
            client_val = row.get(col_client, "")
            if client_val is None:
                client_val = ""
            client_id = normalize_code(client_val)
            
            if not client_id:
                client_id = "Sem ID"
            
            if client_id not in grouped[key]["clients"]:
                grouped[key]["clients"][client_id] = {"client_id": client_id, "recommendations": []}
            
            real_col_ativo_saida = col_ativo_saida or find_column(row, ativo_saida_names)
            real_col_valor_saida = col_valor_saida or find_column(row, valor_saida_names)
            real_col_ativo_compra = col_ativo_compra or find_column(row, ativo_compra_names)
            real_col_valor_compra = col_valor_compra or find_column(row, valor_compra_names)
            
            recommendation = {
                "ativo_saida": str(row.get(real_col_ativo_saida, "") or "") if real_col_ativo_saida else "",
                "valor_saida": format_currency(row.get(real_col_valor_saida, 0)) if real_col_valor_saida else "R$ 0,00",
                "ativo_compra": str(row.get(real_col_ativo_compra, "") or "") if real_col_ativo_compra else "",
                "valor_compra": format_currency(row.get(real_col_valor_compra, 0)) if real_col_valor_compra else "R$ 0,00",
                "client_id": client_id
            }
            
            mapped_currency_cols = set()
            if real_col_valor_saida:
                mapped_currency_cols.add(real_col_valor_saida)
            if real_col_valor_compra:
                mapped_currency_cols.add(real_col_valor_compra)
            
            protected_keys = {"ativo_saida", "valor_saida", "ativo_compra", "valor_compra", "client_id"}
            for col_name, col_val in row.items():
                if col_name in protected_keys:
                    continue
                if col_name in mapped_currency_cols:
                    recommendation[col_name] = format_currency(col_val)
                else:
                    recommendation[col_name] = col_val
            
            grouped[key]["clients"][client_id]["recommendations"].append(recommendation)
            grouped[key]["total_recommendations"] += 1
    
    print(f"[GROUPING] Final result: {len(grouped)} assessors")
    for aid, adata in grouped.items():
        client_count = len(adata.get('clients', {}))
        rec_count = adata.get('total_recommendations', 0)
        print(f"[GROUPING]   {aid}: nome={adata['nome']}, clients={client_count}, recs={rec_count}")
    
    return grouped


def is_currency_value(value) -> bool:
    """Detecta se um valor parece ser monetário (R$ ou número grande)."""
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value >= 100
    if isinstance(value, str):
        s = str(value).strip()
        if "R$" in s:
            return True
        cleaned = s.replace(".", "").replace(",", ".")
        try:
            num = float(cleaned)
            return num >= 100
        except (ValueError, TypeError):
            return False
    return False


def format_currency(value) -> str:
    """Formata valor para moeda brasileira."""
    if value is None:
        return "R$ 0,00"
    if isinstance(value, str) and "R$" in value:
        return value
    try:
        if isinstance(value, str):
            value = value.replace("R$", "").replace(".", "").replace(",", ".").strip()
            if not value:
                return "R$ 0,00"
        num = float(value)
        return f"R$ {num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return str(value) if value else "R$ 0,00"


def build_message(template_content: str, assessor_data: dict, custom_mapping: dict, content_template: str = None, group_by_client: bool = False) -> str:
    """
    Constrói a mensagem final substituindo as variáveis do template.
    
    Args:
        content_template: Template para cada linha de recomendação (bloco repetível).
        group_by_client: Se True, adiciona cabeçalho de cliente no bloco lista_clientes.
    """
    import unicodedata

    def normalize_var_name(name: str) -> str:
        if not name:
            return ""
        normalized = unicodedata.normalize('NFKD', str(name))
        ascii_text = normalized.encode('ASCII', 'ignore').decode('ASCII')
        return ascii_text.lower().replace(' ', '_').replace('-', '_')

    if not template_content:
        template_content = DEFAULT_TEMPLATE_CONTENT
    
    message = str(template_content)
    
    print(f"[BUILD_MSG] Input template (first 200 chars): {message[:200]}")
    if content_template:
        print(f"[BUILD_MSG] Content template for blocks: {content_template[:100]}")
    
    nome = str(assessor_data.get("nome", "") or "")
    primeiro_nome = nome.split()[0] if nome else ""
    
    base_vars = {
        "codigo_ai": str(assessor_data.get("codigo_ai", "") or ""),
        "nome": nome,
        "primeiro_nome": primeiro_nome,
        "email": str(assessor_data.get("email", "") or ""),
        "telefone_whatsapp": str(assessor_data.get("telefone_whatsapp", "") or ""),
        "telefone": str(assessor_data.get("telefone", "") or ""),
        "unidade": str(assessor_data.get("unidade", "") or ""),
        "equipe": str(assessor_data.get("equipe", "") or ""),
        "broker_responsavel": str(assessor_data.get("broker_responsavel", "") or ""),
        "nome_assessor": nome,
        "assessor": nome,
        "assessor_id": str(assessor_data.get("codigo_ai", "") or ""),
    }
    
    for var_name, value in base_vars.items():
        for pattern in [f"{{{{{var_name}}}}}", f"{{{{ {var_name} }}}}", f"{{{var_name}}}"]:
            message = message.replace(pattern, value)
    
    data_atual = datetime.now().strftime("%d/%m/%Y")
    for pattern in ["{{data_atual}}", "{{ data_atual }}", "{data_atual}"]:
        message = message.replace(pattern, data_atual)
    
    spreadsheet_data = assessor_data.get("spreadsheet_data", {})
    if spreadsheet_data:
        print(f"[BUILD_MSG] Spreadsheet data keys: {list(spreadsheet_data.keys())}")
        for col_name, col_value in spreadsheet_data.items():
            val_str = format_cell_value(col_value)
            for pattern in [f"{{{{{col_name}}}}}", f"{{{{ {col_name} }}}}", f"{{{col_name}}}"]:
                message = message.replace(pattern, val_str)
            norm_name = normalize_var_name(col_name)
            if norm_name != col_name:
                for pattern in [f"{{{{{norm_name}}}}}", f"{{{{ {norm_name} }}}}", f"{{{norm_name}}}"]:
                    message = message.replace(pattern, val_str)
    
    custom_fields = assessor_data.get("custom_fields", {})
    for var_name, value in custom_fields.items():
        val_str = format_cell_value(value)
        for pattern in [f"{{{{{var_name}}}}}", f"{{{{ {var_name} }}}}", f"{{{var_name}}}"]:
            message = message.replace(pattern, val_str)
    
    clients = assessor_data.get("clients", {})
    if clients:
        clients_block = build_clients_block(clients, content_template, group_by_client=group_by_client)
        for pattern in ["{{lista_clientes}}", "{{ lista_clientes }}", "{lista_clientes}"]:
            message = message.replace(pattern, clients_block)
    
    message = re.sub(r'\{\{[^}]+\}\}', '', message)
    
    print(f"[BUILD_MSG] Final message (first 300 chars): {message[:300]}")
    
    return message


def build_clients_block(clients: dict, content_template: str = None, group_by_client: bool = True) -> str:
    """
    Constrói o bloco de texto com as recomendações por cliente.
    
    Se group_by_client for True, adiciona cabeçalho "Cliente X" antes de cada grupo.
    Se group_by_client for False, trata como linhas individuais sem cabeçalho de cliente.
    """
    import unicodedata
    
    def normalize_var_name(name: str) -> str:
        """Remove acentos e normaliza nome de variável."""
        if not name:
            return ""
        normalized = unicodedata.normalize('NFKD', str(name))
        ascii_text = normalized.encode('ASCII', 'ignore').decode('ASCII')
        return ascii_text.lower().replace(' ', '_').replace('-', '_')
    
    def replace_vars_in_text(text: str, data: dict) -> str:
        """Substitui variáveis no texto usando os dados fornecidos."""
        if not text:
            return ""
        result = str(text)
        normalized_data = {}
        for key, value in data.items():
            if isinstance(value, (dict, list)):
                continue
            str_value = format_cell_value(value)
            norm_key = normalize_var_name(key)
            normalized_data[norm_key] = str_value
            normalized_data[str(key)] = str_value
        
        pattern = r'\{\{\s*([^}]+?)\s*\}\}'
        def replacer(match):
            var_name = match.group(1).strip()
            normalized_name = normalize_var_name(var_name)
            if normalized_name in normalized_data:
                return normalized_data[normalized_name]
            if var_name in normalized_data:
                return normalized_data[var_name]
            return match.group(0)
        
        result = re.sub(pattern, replacer, result)
        return result
    
    if not clients:
        print("[BUILD_CLIENTS] No clients provided!")
        return "(Nenhuma recomendação encontrada)"
    
    lines = []
    has_custom_template = bool(content_template and content_template.strip())
    
    for client_id, client_data in clients.items():
        if not client_id:
            client_id = "Sem ID"
        
        if group_by_client:
            lines.append(f"Cliente {client_id}")
        
        if isinstance(client_data, dict):
            recommendations = client_data.get("recommendations", [])
        else:
            recommendations = client_data if isinstance(client_data, list) else []
        
        for rec in recommendations:
            if has_custom_template:
                line = replace_vars_in_text(content_template, rec)
                if line.strip():
                    if group_by_client:
                        lines.append(f"- {line}")
                    else:
                        lines.append(line)
            else:
                ativo_saida = rec.get('ativo_saida', 'N/A')
                valor_saida = rec.get('valor_saida', 'R$ 0,00')
                ativo_compra = rec.get('ativo_compra', 'N/A')
                valor_compra = rec.get('valor_compra', 'R$ 0,00')
                
                line = f"- Saia de {valor_saida} em {ativo_saida} e compre {valor_compra} em {ativo_compra}."
                lines.append(line)
        
        if group_by_client:
            lines.append("")
    
    result = "\n".join(lines).strip()
    print(f"[BUILD_CLIENTS] Generated block with {len(lines)} lines for {len(clients)} clients")
    return result


def build_structured_message(
    header: str,
    content_template: str,
    footer: str,
    assessor_data: dict,
    data_rows: list,
    group_by_client: bool = False,
    client_id_column: str = None
) -> str:
    """
    Constrói mensagem estruturada com 3 blocos: cabeçalho, conteúdo repetível, rodapé.
    
    Args:
        header: Texto do cabeçalho (pode conter variáveis do assessor)
        content_template: Template de uma linha de conteúdo (repetido para cada linha/grupo)
        footer: Texto do rodapé (pode conter variáveis do assessor)
        assessor_data: Dados do assessor (codigo_ai, nome, email, etc.)
        data_rows: Lista de linhas de dados do arquivo
        group_by_client: Se True, agrupa linhas por cliente antes de construir
        client_id_column: Nome da coluna para agrupar por cliente
    
    Returns:
        Mensagem final consolidada
    """
    from datetime import datetime
    import unicodedata
    import re
    
    def normalize_var_name(name: str) -> str:
        """Remove acentos e normaliza nome de variável."""
        if not name:
            return ""
        normalized = unicodedata.normalize('NFKD', str(name))
        ascii_text = normalized.encode('ASCII', 'ignore').decode('ASCII')
        return ascii_text.lower().replace(' ', '_').replace('-', '_')
    
    def is_formatted_currency(value) -> bool:
        """Verifica se o valor já está formatado como moeda."""
        if not isinstance(value, str):
            return False
        return value.strip().startswith("R$")
    
    def replace_vars(text: str, vars_dict: dict) -> str:
        """Substitui variáveis no texto, normalizando acentos e priorizando valores formatados."""
        if not text:
            return ""
        result = str(text)
        normalized_data = {}
        for key, value in vars_dict.items():
            if isinstance(value, (dict, list)):
                continue
            norm_key = normalize_var_name(key)
            str_value = format_cell_value(value)
            
            if norm_key in normalized_data:
                existing = normalized_data[norm_key]
                if is_formatted_currency(existing) and not is_formatted_currency(str_value):
                    continue
            
            normalized_data[norm_key] = str_value
            normalized_data[str(key)] = str_value
        
        pattern = r'\{\{\s*([^}]+?)\s*\}\}'
        def replacer(match):
            var_name = match.group(1).strip()
            normalized_name = normalize_var_name(var_name)
            if normalized_name in normalized_data:
                return normalized_data[normalized_name]
            if var_name in normalized_data:
                return normalized_data[var_name]
            return match.group(0)
        
        result = re.sub(pattern, replacer, result)
        return result
    
    base_vars = {
        "codigo_ai": str(assessor_data.get("codigo_ai", "") or ""),
        "nome": str(assessor_data.get("nome", "") or ""),
        "nome_assessor": str(assessor_data.get("nome", "") or ""),
        "assessor": str(assessor_data.get("nome", "") or ""),
        "email": str(assessor_data.get("email", "") or ""),
        "telefone_whatsapp": str(assessor_data.get("telefone_whatsapp", "") or ""),
        "telefone": str(assessor_data.get("telefone", "") or ""),
        "unidade": str(assessor_data.get("unidade", "") or ""),
        "equipe": str(assessor_data.get("equipe", "") or ""),
        "broker_responsavel": str(assessor_data.get("broker_responsavel", "") or ""),
        "data_atual": datetime.now().strftime("%d/%m/%Y"),
    }
    
    header_text = replace_vars(header or "", base_vars)
    footer_text = replace_vars(footer or "", base_vars)
    
    content_lines = []
    
    if group_by_client and client_id_column and data_rows:
        grouped = {}
        for row in data_rows:
            client_id = str(row.get(client_id_column, "Sem ID") or "Sem ID")
            if client_id not in grouped:
                grouped[client_id] = []
            grouped[client_id].append(row)
        
        for client_id, client_rows in grouped.items():
            content_lines.append(f"Cliente {client_id}")
            for row in client_rows:
                row_vars = {**base_vars, **row}
                line = replace_vars(content_template or "", row_vars)
                if line.strip():
                    content_lines.append(f"- {line}")
            content_lines.append("")
    else:
        for row in data_rows:
            row_vars = {**base_vars, **row}
            line = replace_vars(content_template or "", row_vars)
            if line.strip():
                content_lines.append(line)
    
    content_block = "\n".join(content_lines).strip()
    
    message_parts = []
    if header_text.strip():
        message_parts.append(header_text.strip())
    if content_block:
        message_parts.append(content_block)
    if footer_text.strip():
        message_parts.append(footer_text.strip())
    
    final_message = "\n\n".join(message_parts)
    
    final_message = re.sub(r'\{\{[^}]+\}\}', '', final_message)
    
    return final_message


def _batch_resolve_channels(emails: list, db) -> dict:
    """Resolve channel_id Z-API por e-mail de assessor (Task #224).
    Prioridade: Assessor.channel_id direto → UnidadeChannelMapping → None.
    Retorna {email_original: channel_id | None}.
    Normaliza e-mails (lower + strip) para evitar miss por capitalização.
    """
    from database.models import UnidadeChannelMapping
    from sqlalchemy import func as _sa_func
    if not emails:
        return {}

    emails_norm = [e.lower().strip() for e in emails]

    assessors = (
        db.query(Assessor.email, Assessor.channel_id, Assessor.unidade)
        .filter(_sa_func.lower(_sa_func.trim(Assessor.email)).in_(emails_norm))
        .all()
    )

    email_map = {a.email.lower().strip(): (a.channel_id, a.unidade) for a in assessors}

    missing_unidades = {
        u for (_ch, u) in email_map.values()
        if not _ch and u
    }

    unidade_channel = {}
    if missing_unidades:
        for ucm in db.query(UnidadeChannelMapping).filter(
            UnidadeChannelMapping.unidade.in_(missing_unidades)
        ).all():
            unidade_channel[ucm.unidade] = ucm.channel_id

    result = {}
    for email in emails:
        key = email.lower().strip()
        if key in email_map:
            ch, unidade = email_map[key]
            if ch:
                result[email] = ch
            elif unidade and unidade in unidade_channel:
                result[email] = unidade_channel[unidade]
            else:
                result[email] = None
        else:
            result[email] = None

    return result


def _resolve_channel_client_for_dispatch(channel_id, db_session):
    """Task #224 — resolve cliente Z-API para dispatch imediato (não-cadência).

    Returns:
        (client, is_configured, is_inactive)
        - is_inactive=True  → canal inativo; caller deve falhar com "Canal desativado"
        - is_inactive=False, is_configured=True  → client pronto para envio
        - is_inactive=False, is_configured=False → Z-API não configurada; caller pode simular
    """
    from services.whatsapp_client import get_zapi_client_for_channel as _gzc
    from services.whatsapp_client import zapi_client as _legacy

    if channel_id is None:
        return _legacy, _legacy.is_configured(), False

    from database.models import ZAPIChannel as _ZCh
    ch_row = db_session.query(_ZCh).filter(_ZCh.id == channel_id).first()
    if not ch_row or not ch_row.is_active:
        return None, False, True

    client = _gzc(channel_id, db_session)
    return client, client.is_configured(), False


@router.post("/{campaign_id}/dispatch")
async def dispatch_campaign(
    campaign_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    """
    Dispara a campanha enviando mensagens via WhatsApp.
    Task #224 — validação Z-API por canal no loop, sem preflight global legado.
    """
    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campanha não encontrada")
    
    if campaign.status == CampaignStatus.SENT.value:
        raise HTTPException(status_code=400, detail="Esta campanha já foi enviada")
    
    # Check source_type to route to correct dispatch function
    source_type = getattr(campaign, 'source_type', 'upload') or 'upload'
    
    if source_type in ["base", "base_assessores"]:
        return await dispatch_campaign_from_base(campaign, db)
    
    # Usa template customizado se existir, senao template salvo, senao mensagem padrao
    template_content = DEFAULT_TEMPLATE_CONTENT
    
    if campaign.custom_template_content:
        candidate = str(campaign.custom_template_content)
        if template_has_required_variables(candidate):
            template_content = candidate
    elif campaign.template_id:
        template = db.query(MessageTemplate).filter(MessageTemplate.id == campaign.template_id).first()
        if template:
            candidate = str(template.content)
            if template_has_required_variables(candidate):
                template_content = candidate
    
    try:
        column_mapping = json.loads(str(campaign.column_mapping)) if campaign.column_mapping else {}
        custom_mapping = json.loads(str(campaign.custom_fields_mapping)) if campaign.custom_fields_mapping else {}
        data = json.loads(str(campaign.processed_data)) if campaign.processed_data else []
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Erro nos dados da campanha")
    
    grouped = group_recommendations_by_assessor(data, column_mapping, custom_mapping, db)

    # Task #224 — resolve channel_id por assessor em batch (não-SSE dispatch).
    # Preflight global de Z-API removido: validação é por canal no loop.
    from services.whatsapp_client import get_zapi_client_for_channel as _gzc_disp
    from services.whatsapp_client import zapi_client as _legacy_zapi_disp
    _disp_emails = [v.get("email_assessor", "") for v in grouped.values() if v.get("email_assessor")]
    _disp_channel_map = _batch_resolve_channels(_disp_emails, db) if _disp_emails else {}

    # Task #312 — Pre-flight check ANTES de marcar PROCESSING (não-SSE dispatch).
    _pf_disp_cids = list({v for v in _disp_channel_map.values()})
    if not _pf_disp_cids:
        _pf_disp_cids = [None]
    _pf_disp_result = await _run_preflight_check(_pf_disp_cids, db)
    if not _pf_disp_result["all_ok"]:
        _pf_disp_bad = [ch for ch in _pf_disp_result["channels"] if not ch.get("ok")]
        raise HTTPException(
            status_code=422,
            detail={
                "preflight_failed": True,
                "message": "Um ou mais canais de envio apresentam problemas. Corrija antes de disparar.",
                "channels": _pf_disp_bad,
            },
        )

    # Preflight OK — apenas agora a campanha entra em PROCESSING.
    campaign.status = CampaignStatus.PROCESSING.value
    db.commit()

    sent_count = 0
    failed_count = 0
    content_line_template = campaign.message_content_template or ""
    
    is_grouped = bool(campaign.group_by_client)
    for assessor_id, assessor_data in grouped.items():
        message = build_message(template_content, assessor_data, custom_mapping, content_line_template, group_by_client=is_grouped)
        phone = assessor_data.get("telefone", "")
        _a_email = assessor_data.get("email_assessor", "")
        _a_channel_id = _disp_channel_map.get(_a_email)

        dispatch = CampaignDispatch(
            campaign_id=campaign_id,
            assessor_id=assessor_id,
            assessor_email=_a_email,
            assessor_phone=phone,
            assessor_name=assessor_data.get("nome_assessor", ""),
            message_content=message,
            channel_id=_a_channel_id,
            status="pending"
        )
        db.add(dispatch)
        db.flush()

        # Seleciona cliente Z-API por canal (Task #224) — helper compartilhado.
        _active_zapi, _zapi_configured, _chan_inactive_disp = \
            _resolve_channel_client_for_dispatch(_a_channel_id, db)
        if _chan_inactive_disp:
            dispatch.status = "failed"
            dispatch.error_message = "Canal desativado"
            dispatch.error_details = "O canal Z-API associado a este assessor está inativo."
            failed_count += 1
            db.flush()
            continue

        if phone and _zapi_configured:
            try:
                result = await _active_zapi.send_text(phone, message, delay_typing=2)
                dispatch.api_response = json.dumps(result, ensure_ascii=False, default=str)
                
                if result.get("success"):
                    dispatch.status = "sent"
                    dispatch.sent_at = datetime.utcnow()
                    sent_count += 1
                    _persist_campaign_message(db, phone, message, campaign.name, channel_id=_a_channel_id, campaign_id=campaign_id)
                else:
                    dispatch.status = "failed"
                    error_code = result.get("error_code", "UNKNOWN")
                    error_msg = result.get("error", "Erro desconhecido")
                    dispatch.error_message = error_msg
                    dispatch.error_details = translate_error_to_natural_language(error_code, error_msg, phone)
                    failed_count += 1
                    print(f"[DISPATCH-FAIL] canal={_a_channel_id} assessor={_a_email} motivo={error_code} detalhe={error_msg[:200]} api_response={str(result)[:300]}")
            except Exception as e:
                dispatch.status = "failed"
                dispatch.error_message = str(e)
                dispatch.error_details = f"Erro inesperado ao enviar mensagem: {str(e)}"
                failed_count += 1
        else:
            if not phone:
                dispatch.status = "failed"
                dispatch.error_message = "Telefone não informado"
                dispatch.error_details = f"O assessor '{assessor_data.get('nome_assessor', 'Desconhecido')}' não possui número de telefone cadastrado na planilha ou na base de assessores."
                failed_count += 1
            elif not _zapi_configured:
                dispatch.status = "simulated"
                dispatch.error_details = "Disparo simulado - Z-API não configurado"
                dispatch.sent_at = datetime.utcnow()
                sent_count += 1
    
    campaign.status = CampaignStatus.SENT.value
    campaign.messages_sent = sent_count
    campaign.messages_failed = failed_count
    campaign.sent_at = datetime.utcnow()
    campaign.total_assessors = len(grouped)
    db.commit()
    
    return {
        "message": "Campanha disparada com sucesso",
        "total_assessors": len(grouped),
        "messages_sent": sent_count,
        "messages_failed": failed_count
    }


async def _run_preflight_check(channel_ids_raw: list, db) -> dict:
    """
    Task #312 — Verifica pré-condições para disparo por canal:
    (1) Conectividade da instância Z-API.
    (2) Webhook de recebimento configurado (apenas canais não-legados).

    - channel_ids_raw: lista de channel_id (pode conter None = canal legado).
    - Retorna { channels: [...], all_ok: bool }.
    - Nunca lança exceção — erros de verificação são tratados como aviso;
      falhas de rede no check do webhook são fail-safe (não bloqueiam).
    """
    from services.dependency_check import (
        get_channel_health_cache,
        get_zapi_status_cache,
        check_zapi_connectivity,
        check_zapi_connectivity_for_channel,
    )
    from services.whatsapp_client import get_zapi_client_for_channel as _gzc_pf
    from database.models import ZAPIChannel as _ZCh_pf
    from core.config import get_public_base_url as _gpbu
    import os as _os_pf

    public_base = _gpbu()
    results: list = []
    all_ok = True
    seen: set = set()

    for channel_id in channel_ids_raw:
        cid_key = channel_id if channel_id is not None else "legacy"
        if cid_key in seen:
            continue
        seen.add(cid_key)

        # ── Canal legado (None ou is_legacy=True identificado mais tarde) ──────
        if channel_id is None:
            legacy_cache = get_zapi_status_cache()
            status = legacy_cache.get("status", "unknown")
            connectivity_ok = status == "connected"

            if not connectivity_ok and status in ("unknown", "error", "timeout"):
                try:
                    direct = await check_zapi_connectivity()
                    status = direct.get("status", "unknown")
                    connectivity_ok = status == "connected"
                except Exception:
                    pass

            results.append({
                "channel_id": None,
                "label": "Canal legado",
                "is_legacy": True,
                "connectivity_ok": connectivity_ok,
                "webhook_ok": True,
                "ok": connectivity_ok,
                "error": (
                    None if connectivity_ok
                    else f"Canal legado — instância desconectada (status: {status}). "
                         "Reconecte no painel Z-API antes de disparar."
                ),
            })
            if not connectivity_ok:
                all_ok = False
            continue

        # ── Canal explícito ──────────────────────────────────────────────────────
        try:
            ch_row = db.query(_ZCh_pf).filter(_ZCh_pf.id == channel_id).first()
        except Exception as _qe:
            results.append({
                "channel_id": channel_id,
                "label": f"Canal #{channel_id}",
                "is_legacy": False,
                "connectivity_ok": False,
                "webhook_ok": False,
                "ok": False,
                "error": f"Canal #{channel_id} — erro ao consultar banco de dados: {type(_qe).__name__}.",
            })
            all_ok = False
            continue

        if not ch_row:
            results.append({
                "channel_id": channel_id,
                "label": f"Canal #{channel_id}",
                "is_legacy": False,
                "connectivity_ok": False,
                "webhook_ok": False,
                "ok": False,
                "error": f"Canal #{channel_id} — não encontrado na base de dados.",
            })
            all_ok = False
            continue

        label = ch_row.label or ch_row.name or f"Canal #{channel_id}"

        if not ch_row.is_active:
            results.append({
                "channel_id": channel_id,
                "label": label,
                "is_legacy": bool(ch_row.is_legacy),
                "connectivity_ok": False,
                "webhook_ok": False,
                "ok": False,
                "error": f"{label} — canal desativado. Ative em Integrações → Canais antes de disparar.",
            })
            all_ok = False
            continue

        # Canais legados identificados na tabela: verifica apenas conectividade via cache global
        if ch_row.is_legacy:
            legacy_cache = get_zapi_status_cache()
            status = legacy_cache.get("status", "unknown")
            connectivity_ok = status == "connected"
            results.append({
                "channel_id": channel_id,
                "label": label,
                "is_legacy": True,
                "connectivity_ok": connectivity_ok,
                "webhook_ok": True,
                "ok": connectivity_ok,
                "error": (
                    None if connectivity_ok
                    else f"{label} — instância desconectada (status: {status}). "
                         "Reconecte no painel Z-API antes de disparar."
                ),
            })
            if not connectivity_ok:
                all_ok = False
            continue

        # Canal explícito não-legado ─ verifica conectividade + webhook
        # 1. Conectividade: cache Task #308 com fallback para chamada direta
        health_cache = get_channel_health_cache()
        cached = health_cache.get(channel_id)
        connectivity_status = cached.get("status", "unknown") if cached else "unknown"

        if not cached or connectivity_status in ("unknown", "error"):
            try:
                global_ct = _os_pf.getenv("ZAPI_CLIENT_TOKEN", "")
                direct = await check_zapi_connectivity_for_channel(
                    ch_row.instance_id, ch_row.token, ch_row.client_token, global_ct
                )
                connectivity_status = direct.get("status", "unknown")
            except Exception:
                pass

        connectivity_ok = connectivity_status == "connected"

        # 2. Webhook: apenas canais não-legados (Task #309 mantém registro; aqui verificamos)
        webhook_ok = True
        webhook_error_detail = None
        try:
            _pf_client = _gzc_pf(channel_id, db)
            webhook_resp = await _pf_client.get_webhook_settings(timeout=4.0)
            if webhook_resp.get("endpoint_not_found"):
                webhook_ok = True  # endpoint não suportado — não bloqueia
            elif webhook_resp.get("success"):
                settings = webhook_resp.get("settings") or {}
                received_url = ""
                if isinstance(settings, dict):
                    # Formato 1: {"webhookReceived": {"value": "..."}}
                    wr = settings.get("webhookReceived") or {}
                    if isinstance(wr, dict):
                        received_url = str(wr.get("value", "") or "")
                    # Formato 2: {"received": {"value": "..."}}
                    if not received_url:
                        wr2 = settings.get("received") or {}
                        if isinstance(wr2, dict):
                            received_url = str(wr2.get("value", "") or "")
                    # Formato plano: {"value": "..."}
                    if not received_url:
                        received_url = str(settings.get("value", "") or "")
                received_url = received_url.strip()

                if not received_url:
                    webhook_ok = False
                    webhook_error_detail = "webhook de recebimento não configurado"
                elif public_base:
                    expected_path = f"/api/whatsapp/webhook/{channel_id}"
                    if public_base not in received_url and expected_path not in received_url:
                        webhook_ok = False
                        webhook_error_detail = (
                            f"webhook aponta para URL diferente da esperada "
                            f"({received_url[:60]}{'…' if len(received_url) > 60 else ''})"
                        )
                # URL configurada mas base URL desconhecida → não dá pra validar destino; ok
            else:
                # Resposta não-success e não endpoint_not_found (ex: credenciais inválidas,
                # instância temporariamente em erro). Política explícita: fail-safe —
                # não bloqueia o disparo. A verificação de conectividade (separada acima)
                # já captura instâncias verdadeiramente offline. Impedir disparo aqui
                # criaria falsos-positivos em instâncias lentas ou temporariamente
                # inacessíveis pelo settings endpoint.
                webhook_ok = True
        except Exception:
            # Erro transitório de rede (timeout, DNS, etc.) — a chamada não completou.
            # Diferente do else acima (resposta recebida mas com erro): aqui não há dados.
            # Política: fail-safe — não bloqueia. Trata-se de indisponibilidade transitória,
            # não de evidência de má-configuração.
            webhook_ok = True

        ok = connectivity_ok and webhook_ok
        error_parts = []
        if not connectivity_ok:
            error_parts.append(
                f"instância desconectada (status: {connectivity_status}). "
                "Reconecte no painel Z-API antes de disparar."
            )
        if not webhook_ok and webhook_error_detail:
            error_parts.append(f"{webhook_error_detail}. Configure em Integrações → Canais.")

        results.append({
            "channel_id": channel_id,
            "label": label,
            "is_legacy": False,
            "connectivity_ok": connectivity_ok,
            "webhook_ok": webhook_ok,
            "ok": ok,
            "error": f"{label} — " + " | ".join(error_parts) if error_parts else None,
        })
        if not ok:
            all_ok = False

    return {"channels": results, "all_ok": all_ok}


@router.post("/{campaign_id}/preflight-check")
async def preflight_campaign_channels(
    campaign_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao()),
):
    """
    Task #312 — Verifica pré-condições de disparo (conectividade + webhook) para todos
    os canais que serão utilizados nesta campanha.

    Resolve os channel_ids a partir de `processed_data` (suporta upload e base)
    e delega para `_run_preflight_check`. Retorna { channels: [...], all_ok: bool }.
    Nunca retorna 422 — é uma consulta informacional; quem decide bloquear é o caller.
    """
    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campanha não encontrada")

    try:
        proc = json.loads(str(campaign.processed_data)) if campaign.processed_data else []
    except Exception:
        proc = []

    emails: list = []
    for row in proc:
        e = (
            row.get("email_assessor")
            or row.get("email")
            or row.get("Email")
            or row.get("assessor_email")
            or ""
        )
        if e:
            emails.append(str(e).strip().lower())
    emails = [e for e in emails if e]

    channel_map = _batch_resolve_channels(list(set(emails)), db) if emails else {}
    channel_ids: list = list({v for v in channel_map.values()})
    if not channel_ids:
        channel_ids = [None]

    result = await _run_preflight_check(channel_ids, db)
    return result


@router.get("/{campaign_id}/dispatch-stream")
async def dispatch_campaign_stream(
    campaign_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    """
    Dispara a campanha com streaming de progresso via SSE.
    Envia mensagens uma a uma com delay para evitar sobrecarga.
    Task #224 — valida Z-API por canal no loop, não globalmente.
    """
    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campanha não encontrada")
    
    if campaign.status == CampaignStatus.SENT.value:
        raise HTTPException(status_code=400, detail="Esta campanha já foi enviada")
    
    source_type = getattr(campaign, 'source_type', 'upload') or 'upload'
    
    if source_type in ["base", "base_assessores"]:
        return await dispatch_campaign_from_base(campaign, db)
    
    header_template = campaign.message_header or ""
    content_template = campaign.message_content_template or ""
    footer_template = campaign.message_footer or ""
    use_blocks = bool(header_template.strip() or content_template.strip() or footer_template.strip())
    
    template_content = DEFAULT_TEMPLATE_CONTENT
    
    if not use_blocks:
        if campaign.custom_template_content:
            candidate = str(campaign.custom_template_content)
            if template_has_required_variables(candidate):
                template_content = candidate
        elif campaign.template_id:
            template = db.query(MessageTemplate).filter(MessageTemplate.id == campaign.template_id).first()
            if template:
                candidate = str(template.content)
                if template_has_required_variables(candidate):
                    template_content = candidate
    
    try:
        column_mapping = json.loads(str(campaign.column_mapping)) if campaign.column_mapping else {}
        custom_mapping = json.loads(str(campaign.custom_fields_mapping)) if campaign.custom_fields_mapping else {}
        data = json.loads(str(campaign.processed_data)) if campaign.processed_data else []
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Erro nos dados da campanha")
    
    grouped = group_recommendations_by_assessor(data, column_mapping, custom_mapping, db)
    total_assessors = len(grouped)
    
    if total_assessors == 0:
        campaign.status = CampaignStatus.SENT.value
        campaign.messages_sent = 0
        campaign.messages_failed = 0
        campaign.sent_at = datetime.utcnow()
        campaign.total_assessors = 0
        db.commit()
        
        async def empty_generator():
            yield f"data: {json.dumps({'type': 'start', 'total': 0})}\n\n"
            yield f"data: {json.dumps({'type': 'complete', 'total': 0, 'sent_count': 0, 'failed_count': 0})}\n\n"
        
        return StreamingResponse(
            empty_generator(),
            media_type="text/event-stream",
            headers={"Cache-Control": "no-cache", "Connection": "keep-alive", "X-Accel-Buffering": "no"}
        )
    
    attachment_url = campaign.attachment_url
    attachment_type = campaign.attachment_type
    attachment_filename = campaign.attachment_filename
    
    content_line_template = campaign.message_content_template or ""
    is_grouped = bool(campaign.group_by_client)

    # Task #224 — pré-resolver channel_id por e-mail dos assessores desta campanha.
    _sse_all_emails = [v.get("email_assessor", "") for v in grouped.values() if v.get("email_assessor")]
    _sse_channel_map = _batch_resolve_channels(_sse_all_emails, db) if _sse_all_emails else {}

    # Task #312 — Pre-flight check ANTES de marcar PROCESSING (SSE upload path).
    _pf_sse_cids = list({v for v in _sse_channel_map.values()})
    if not _pf_sse_cids:
        _pf_sse_cids = [None]
    _pf_sse_result = await _run_preflight_check(_pf_sse_cids, db)
    if not _pf_sse_result["all_ok"]:
        _pf_sse_bad = [ch for ch in _pf_sse_result["channels"] if not ch.get("ok")]
        raise HTTPException(
            status_code=422,
            detail={
                "preflight_failed": True,
                "message": "Um ou mais canais de envio apresentam problemas. Corrija antes de disparar.",
                "channels": _pf_sse_bad,
            },
        )

    # Preflight OK — apenas agora a campanha entra em PROCESSING.
    campaign.status = CampaignStatus.PROCESSING.value
    campaign.total_assessors = total_assessors
    db.commit()

    async def generate_events():
        from core.config import resolve_attachment_for_send
        import os
        
        # Task #224 — sem zapi_configured global; validação é per-dispatch via helper.
        # Resolve o anexo UMA VEZ por campanha — preferindo base64 quando o
        # arquivo existe localmente (elimina problema de URL inacessível pelo
        # Z-API, como janeway.replit.dev). Fallback para URL pública se o
        # arquivo não existir no disco.
        full_attachment_url = resolve_attachment_for_send(attachment_url) if attachment_url else None
        attachment_url_invalid = bool(attachment_url) and full_attachment_url is None
        sent_count = 0
        failed_count = 0
        current_index = 0
        cancelled = False
        
        try:
            yield f"data: {json.dumps({'type': 'start', 'total': total_assessors})}\n\n"
            
            for assessor_id, assessor_data in grouped.items():
                current_index += 1
                
                if use_blocks:
                    wrapper_parts = []
                    if header_template.strip():
                        wrapper_parts.append(header_template.strip())
                    
                    if content_template.strip() and "{{lista_clientes}}" in content_template:
                        wrapper_parts.append(content_template.strip())
                    else:
                        wrapper_parts.append("{{lista_clientes}}")
                    
                    if footer_template.strip():
                        wrapper_parts.append(footer_template.strip())
                    
                    wrapper_template = "\n\n".join(wrapper_parts)
                    message = build_message(wrapper_template, assessor_data, custom_mapping, content_line_template, group_by_client=is_grouped)
                else:
                    message = build_message(template_content, assessor_data, custom_mapping, content_line_template, group_by_client=is_grouped)
                
                phone = assessor_data.get("telefone", "")
                assessor_name = assessor_data.get("nome_assessor", "")
                
                db_session = SessionLocal()
                try:
                    # Task #224 — resolve channel_id para este assessor.
                    _sse_channel_id = _sse_channel_map.get(assessor_data.get("email_assessor", ""))
                    dispatch = CampaignDispatch(
                        campaign_id=campaign_id,
                        assessor_id=assessor_id,
                        assessor_email=assessor_data.get("email_assessor", ""),
                        assessor_phone=phone,
                        assessor_name=assessor_name,
                        message_content=message,
                        status="pending",
                        channel_id=_sse_channel_id,
                    )
                    db_session.add(dispatch)
                    db_session.flush()

                    # Task #224 — resolve cliente por canal via helper compartilhado.
                    _active_client, _act_cfg, _chan_inactive_sse = \
                        _resolve_channel_client_for_dispatch(_sse_channel_id, db_session)
                    if _chan_inactive_sse:
                        dispatch.status = "failed"
                        dispatch.error_message = "Canal desativado"
                        dispatch.error_details = (
                            "O canal Z-API associado a este assessor está "
                            "inativo. Ative o canal em Integrações → Canais."
                        )
                        failed_count += 1
                        status = "failed"
                        error_msg = "Canal desativado"

                    if not _chan_inactive_sse:
                        status = "pending"
                        error_msg = ""
                    attempt = 1

                    if not _chan_inactive_sse and phone and _act_cfg and attachment_url_invalid:
                        # Anexo configurado mas URL pública não pôde ser
                        # construída (sem APP_BASE_URL/REPLIT_DOMAINS).
                        # Mandar caminho relativo para o Z-API faz o disparo
                        # travar em "pendente" eternamente. Falhar agora
                        # com mensagem clara.
                        dispatch.status = "failed"
                        dispatch.error_message = "Arquivo do anexo não encontrado"
                        dispatch.error_details = (
                            "O arquivo do anexo não pôde ser resolvido para envio "
                            "via WhatsApp. Causas possíveis: (1) arquivo não "
                            "encontrado no servidor — verifique se o upload foi "
                            "feito no ambiente de produção (não no ambiente de dev); "
                            "(2) variável APP_BASE_URL não configurada no Railway — "
                            "configure com o domínio público da aplicação "
                            "(ex.: https://agente-ia-rv.railway.app)."
                        )
                        failed_count += 1
                        status = "failed"
                        error_msg = "Arquivo do anexo não encontrado"
                    elif not _chan_inactive_sse and phone and _act_cfg:
                        while attempt <= MAX_RETRY_ATTEMPTS:
                            try:
                                if attachment_url and attachment_type:
                                    if attachment_type == "image":
                                        result = await _active_client.send_image(phone, full_attachment_url, message)
                                    elif attachment_type == "video":
                                        result = await _active_client.send_video(phone, full_attachment_url, message)
                                    elif attachment_type == "audio":
                                        result = await _active_client.send_audio(phone, full_attachment_url)
                                    else:
                                        result = await _active_client.send_document(phone, full_attachment_url, attachment_filename or "", message)
                                else:
                                    result = await _active_client.send_text(phone, message, delay_typing=2)
                                dispatch.api_response = json.dumps(result, ensure_ascii=False, default=str)
                                
                                if result.get("success"):
                                    dispatch.status = "sent"
                                    dispatch.sent_at = datetime.utcnow()
                                    sent_count += 1
                                    status = "sent"
                                    _persist_campaign_message(db_session, phone, message, campaign.name, channel_id=_sse_channel_id, campaign_id=campaign_id)
                                    break
                                else:
                                    error_code = result.get("error_code", "UNKNOWN")
                                    error_msg = result.get("error", "Erro desconhecido")
                                    print(f"[DISPATCH-FAIL] canal={_sse_channel_id} assessor={assessor_data.get('email_assessor','')} motivo={error_code} detalhe={str(error_msg)[:200]} api_response={str(result)[:300]}")
                                    
                                    is_retryable = (
                                        error_code.startswith("HTTP_5") or 
                                        "500" in error_code or
                                        "502" in error_code or
                                        "503" in error_code or
                                        error_code in ["TIMEOUT", "CONNECTION_ERROR", "HTTP_ERROR"]
                                    )
                                    
                                    if is_retryable and attempt < MAX_RETRY_ATTEMPTS:
                                        retry_data = {
                                            'type': 'retry',
                                            'current': current_index,
                                            'total': total_assessors,
                                            'assessor_name': assessor_name,
                                            'attempt': attempt,
                                            'max_attempts': MAX_RETRY_ATTEMPTS,
                                            'error': error_msg
                                        }
                                        yield f"data: {json.dumps(retry_data, ensure_ascii=False)}\n\n"
                                        await asyncio.sleep(RETRY_DELAY_SECONDS)
                                        attempt += 1
                                        continue
                                    else:
                                        dispatch.status = "failed"
                                        dispatch.error_message = error_msg
                                        dispatch.error_details = translate_error_to_natural_language(error_code, error_msg, phone)
                                        if attempt > 1:
                                            dispatch.error_details += f" (após {attempt} tentativas)"
                                        failed_count += 1
                                        status = "failed"
                                        break
                            except Exception as e:
                                error_msg = str(e)
                                
                                if attempt < MAX_RETRY_ATTEMPTS:
                                    retry_data = {
                                        'type': 'retry',
                                        'current': current_index,
                                        'total': total_assessors,
                                        'assessor_name': assessor_name,
                                        'attempt': attempt,
                                        'max_attempts': MAX_RETRY_ATTEMPTS,
                                        'error': error_msg
                                    }
                                    yield f"data: {json.dumps(retry_data, ensure_ascii=False)}\n\n"
                                    await asyncio.sleep(RETRY_DELAY_SECONDS)
                                    attempt += 1
                                    continue
                                else:
                                    dispatch.status = "failed"
                                    dispatch.error_message = error_msg
                                    dispatch.error_details = f"Erro inesperado ao enviar mensagem: {error_msg}"
                                    if attempt > 1:
                                        dispatch.error_details += f" (após {attempt} tentativas)"
                                    failed_count += 1
                                    status = "failed"
                                    break
                    elif not _chan_inactive_sse:
                        if not phone:
                            dispatch.status = "failed"
                            dispatch.error_message = "Telefone não informado"
                            dispatch.error_details = f"O assessor '{assessor_name}' não possui número de telefone cadastrado."
                            failed_count += 1
                            status = "failed"
                            error_msg = "Telefone não informado"
                        elif not _act_cfg:
                            dispatch.status = "simulated"
                            dispatch.error_details = "Disparo simulado - Z-API não configurado"
                            dispatch.sent_at = datetime.utcnow()
                            sent_count += 1
                            status = "simulated"
                    
                    db_session.commit()
                    
                    percent = round((current_index / total_assessors) * 100, 1)
                    progress_data = {
                        'type': 'progress',
                        'current': current_index,
                        'total': total_assessors,
                        'percent': percent,
                        'assessor_name': assessor_name,
                        'assessor_phone': phone,
                        'status': status,
                        'error': error_msg,
                        'sent_count': sent_count,
                        'failed_count': failed_count,
                        'attempts_made': attempt
                    }
                    yield f"data: {json.dumps(progress_data, ensure_ascii=False)}\n\n"
                    
                finally:
                    db_session.close()
                
                if current_index < total_assessors:
                    delay = get_random_dispatch_delay()
                    await asyncio.sleep(delay)
        
        except asyncio.CancelledError:
            cancelled = True
        finally:
            db_final = SessionLocal()
            try:
                campaign_final = db_final.query(Campaign).filter(Campaign.id == campaign_id).first()
                if campaign_final:
                    campaign_final.status = CampaignStatus.SENT.value
                    campaign_final.messages_sent = sent_count
                    campaign_final.messages_failed = failed_count
                    campaign_final.sent_at = datetime.utcnow()
                    db_final.commit()
            finally:
                db_final.close()
        
        if not cancelled:
            complete_data = {
                'type': 'complete',
                'total': total_assessors,
                'sent_count': sent_count,
                'failed_count': failed_count
            }
            yield f"data: {json.dumps(complete_data)}\n\n"
    
    return StreamingResponse(
        generate_events(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"
        }
    )


def replace_variables_generic(text: str, variables: dict) -> str:
    """
    Substitui variáveis no texto de forma genérica.
    Suporta formatos: {{var}}, {{ var }}, {var}
    Normaliza nomes de variáveis (remove acentos, lowercase, underscores).
    """
    if not text:
        return ""
    
    def normalize_var_name(name: str) -> str:
        import unicodedata
        normalized = unicodedata.normalize('NFKD', name.lower())
        normalized = ''.join(c for c in normalized if not unicodedata.combining(c))
        normalized = normalized.replace(' ', '_').replace('-', '_')
        return normalized
    
    normalized_vars = {}
    for key, value in variables.items():
        str_value = format_cell_value(value)
        normalized_vars[key] = str_value
        normalized_vars[normalize_var_name(key)] = str_value
    
    result = text
    
    pattern = r'\{\{\s*([^}]+?)\s*\}\}'
    def replacer(match):
        var_name = match.group(1).strip()
        normalized_name = normalize_var_name(var_name)
        if var_name in normalized_vars:
            return normalized_vars[var_name]
        if normalized_name in normalized_vars:
            return normalized_vars[normalized_name]
        return match.group(0)
    
    result = re.sub(pattern, replacer, result)
    
    pattern_simple = r'\{([^{}]+)\}'
    def replacer_simple(match):
        var_name = match.group(1).strip()
        normalized_name = normalize_var_name(var_name)
        if var_name in normalized_vars:
            return normalized_vars[var_name]
        if normalized_name in normalized_vars:
            return normalized_vars[normalized_name]
        return match.group(0)
    
    result = re.sub(pattern_simple, replacer_simple, result)
    
    return result


def build_assessor_variables(assessor: dict) -> dict:
    """
    Constrói dicionário completo de variáveis a partir dos dados do assessor.
    Inclui aliases comuns para facilitar o uso de variáveis.
    """
    variables = {}
    
    print(f"[BUILD_VARS] Input assessor keys: {list(assessor.keys())}")
    print(f"[BUILD_VARS] nome value: {assessor.get('nome', 'NOT_FOUND')}")
    
    for key, value in assessor.items():
        str_value = format_cell_value(value)
        variables[key] = str_value
    
    nome = assessor.get("nome", "")
    primeiro_nome = str(nome).split()[0] if nome else ""
    variables["primeiro_nome"] = primeiro_nome
    variables["nome_assessor"] = str(nome) if nome else ""
    variables["assessor"] = str(nome) if nome else ""
    
    print(f"[BUILD_VARS] primeiro_nome calculated: '{primeiro_nome}' from nome: '{nome}'")
    
    telefone = assessor.get("telefone_whatsapp", "") or assessor.get("telefone", "")
    variables["telefone"] = str(telefone) if telefone else ""
    variables["whatsapp"] = str(telefone) if telefone else ""
    variables["celular"] = str(telefone) if telefone else ""
    
    email = assessor.get("email", "")
    variables["email_assessor"] = str(email) if email else ""
    
    codigo = assessor.get("codigo_ai", "")
    variables["codigo"] = str(codigo) if codigo else ""
    variables["codigo_assessor"] = str(codigo) if codigo else ""
    
    unidade = assessor.get("unidade", "")
    variables["escritorio"] = str(unidade) if unidade else ""
    
    equipe = assessor.get("equipe", "")
    variables["time"] = str(equipe) if equipe else ""
    
    broker = assessor.get("broker_responsavel", "")
    variables["broker"] = str(broker) if broker else ""
    
    variables["data_atual"] = datetime.now().strftime("%d/%m/%Y")
    variables["data"] = datetime.now().strftime("%d/%m/%Y")
    
    variables["lista_clientes"] = "(Campanha informativa)"
    
    return variables


async def dispatch_campaign_from_base(campaign, db: Session):
    """
    Dispara campanha baseada em assessores selecionados da base.
    Envia mensagem usando os blocos de header, content e footer definidos na campanha.
    Task #224 — validação Z-API por canal no loop, sem preflight global legado.
    """
    import os
    
    header_template = campaign.message_header or ""
    content_template = campaign.message_content_template or ""
    footer_template = campaign.message_footer or ""
    
    if not header_template and not content_template and not footer_template:
        if campaign.custom_template_content:
            content_template = str(campaign.custom_template_content)
        elif campaign.template_id:
            template = db.query(MessageTemplate).filter(MessageTemplate.id == campaign.template_id).first()
            if template:
                content_template = str(template.content)
        else:
            content_template = "Ola, {{nome_assessor}}!"
    
    try:
        data = json.loads(str(campaign.processed_data)) if campaign.processed_data else []
    except json.JSONDecodeError:
        data = []
    
    total_assessors = len(data)
    
    if total_assessors == 0:
        campaign.status = CampaignStatus.SENT.value
        campaign.messages_sent = 0
        campaign.messages_failed = 0
        campaign.sent_at = datetime.utcnow()
        campaign.total_assessors = 0
        db.commit()
        
        async def empty_generator():
            yield f"data: {json.dumps({'type': 'start', 'total': 0})}\n\n"
            yield f"data: {json.dumps({'type': 'complete', 'total': 0, 'sent_count': 0, 'failed_count': 0})}\n\n"
        
        return StreamingResponse(
            empty_generator(),
            media_type="text/event-stream",
            headers={"Cache-Control": "no-cache", "Connection": "keep-alive", "X-Accel-Buffering": "no"}
        )
    
    # Task #224 — pré-resolver channel_id para cada assessor da base.
    _base_emails = [a.get("email", "") for a in data if a.get("email")]
    _base_channel_map = _batch_resolve_channels(_base_emails, db) if _base_emails else {}

    # Task #312 — Pre-flight check ANTES de marcar PROCESSING (base path).
    _pf_base_cids = list({v for v in _base_channel_map.values()})
    if not _pf_base_cids:
        _pf_base_cids = [None]
    _pf_base_result = await _run_preflight_check(_pf_base_cids, db)
    if not _pf_base_result["all_ok"]:
        _pf_base_bad = [ch for ch in _pf_base_result["channels"] if not ch.get("ok")]
        raise HTTPException(
            status_code=422,
            detail={
                "preflight_failed": True,
                "message": "Um ou mais canais de envio apresentam problemas. Corrija antes de disparar.",
                "channels": _pf_base_bad,
            },
        )

    # Preflight OK — apenas agora a campanha entra em PROCESSING.
    campaign.status = CampaignStatus.PROCESSING.value
    campaign.total_assessors = total_assessors
    db.commit()

    attachment_url = campaign.attachment_url
    attachment_type = campaign.attachment_type
    attachment_filename = campaign.attachment_filename
    
    async def generate_events():
        from core.config import resolve_attachment_for_send
        # Task #224 — sem zapi_configured global; validação é per-dispatch via helper.
        # Resolve o anexo UMA VEZ por campanha — preferindo base64 quando o
        # arquivo existe localmente (elimina problema de URL inacessível pelo
        # Z-API, como janeway.replit.dev). Fallback para URL pública se o
        # arquivo não existir no disco.
        full_attachment_url = resolve_attachment_for_send(attachment_url) if attachment_url else None
        attachment_url_invalid = bool(attachment_url) and full_attachment_url is None
        sent_count = 0
        failed_count = 0
        current_index = 0
        cancelled = False
        
        try:
            yield f"data: {json.dumps({'type': 'start', 'total': total_assessors})}\n\n"
            
            for assessor in data:
                db_check = SessionLocal()
                try:
                    campaign_check = db_check.query(Campaign).filter(Campaign.id == campaign.id).first()
                    should_cancel = (
                        cancelled_campaigns.get(campaign.id, False) or 
                        (campaign_check and campaign_check.status in ["cancelling", "cancelled"])
                    )
                    if should_cancel:
                        cancelled = True
                        cancel_data = {
                            'type': 'cancelled',
                            'current': current_index,
                            'total': total_assessors,
                            'sent_count': sent_count,
                            'failed_count': failed_count,
                            'message': 'Envio cancelado pelo usuário'
                        }
                        yield f"data: {json.dumps(cancel_data, ensure_ascii=False)}\n\n"
                        break
                finally:
                    db_check.close()
                
                current_index += 1
                assessor_name = assessor.get("nome", "")
                phone = assessor.get("telefone_whatsapp", "") or assessor.get("telefone", "")
                
                variables = build_assessor_variables(assessor)
                
                message_parts = []
                
                header_rendered = replace_variables_generic(header_template, variables)
                if header_rendered.strip():
                    message_parts.append(header_rendered.strip())
                
                content_rendered = replace_variables_generic(content_template, variables)
                if content_rendered.strip():
                    message_parts.append(content_rendered.strip())
                
                footer_rendered = replace_variables_generic(footer_template, variables)
                if footer_rendered.strip():
                    message_parts.append(footer_rendered.strip())
                
                message = "\n\n".join(message_parts)
                
                leftover_pattern = r'\{\{[^}]+\}\}'
                message = re.sub(leftover_pattern, '', message)
                
                db_session = SessionLocal()
                try:
                    # Task #224 — channel_id por e-mail do assessor.
                    _base_channel_id = _base_channel_map.get(assessor.get("email", ""))
                    dispatch = CampaignDispatch(
                        campaign_id=campaign.id,
                        assessor_id=str(assessor.get("id", "")),
                        assessor_email=assessor.get("email", ""),
                        assessor_phone=phone,
                        assessor_name=assessor_name,
                        message_content=message,
                        status="pending",
                        channel_id=_base_channel_id,
                    )
                    db_session.add(dispatch)
                    db_session.flush()

                    # Task #224 — resolve cliente por canal via helper compartilhado.
                    _base_active_client, _base_act_cfg, _chan_inactive_base = \
                        _resolve_channel_client_for_dispatch(_base_channel_id, db_session)
                    if _chan_inactive_base:
                        dispatch.status = "failed"
                        dispatch.error_message = "Canal desativado"
                        dispatch.error_details = (
                            "O canal Z-API associado a este assessor está "
                            "inativo. Ative o canal em Integrações → Canais."
                        )
                        failed_count += 1
                        status = "failed"
                        error_msg = "Canal desativado"

                    if not _chan_inactive_base:
                        status = "pending"
                        error_msg = ""
                    attempt = 1

                    if not _chan_inactive_base and phone and _base_act_cfg and attachment_url_invalid:
                        # Anexo configurado mas URL pública não pôde ser
                        # construída (sem APP_BASE_URL/REPLIT_DOMAINS).
                        # Mandar caminho relativo para o Z-API faz o disparo
                        # travar em "pendente" eternamente. Falhar agora
                        # com mensagem clara.
                        dispatch.status = "failed"
                        dispatch.error_message = "Arquivo do anexo não encontrado"
                        dispatch.error_details = (
                            "O arquivo do anexo não pôde ser resolvido para envio "
                            "via WhatsApp. Causas possíveis: (1) arquivo não "
                            "encontrado no servidor — verifique se o upload foi "
                            "feito no ambiente de produção (não no ambiente de dev); "
                            "(2) variável APP_BASE_URL não configurada no Railway — "
                            "configure com o domínio público da aplicação "
                            "(ex.: https://agente-ia-rv.railway.app)."
                        )
                        failed_count += 1
                        status = "failed"
                        error_msg = "Arquivo do anexo não encontrado"
                    elif not _chan_inactive_base and phone and _base_act_cfg:
                        while attempt <= MAX_RETRY_ATTEMPTS:
                            try:
                                if attachment_url and attachment_type:
                                    if attachment_type == "image":
                                        result = await _base_active_client.send_image(phone, full_attachment_url, message)
                                    elif attachment_type == "video":
                                        result = await _base_active_client.send_video(phone, full_attachment_url, message)
                                    elif attachment_type == "audio":
                                        result = await _base_active_client.send_audio(phone, full_attachment_url)
                                    else:
                                        result = await _base_active_client.send_document(phone, full_attachment_url, attachment_filename or "", message)
                                else:
                                    result = await _base_active_client.send_text(phone, message, delay_typing=2)
                                dispatch.api_response = json.dumps(result, ensure_ascii=False, default=str)
                                
                                if result.get("success"):
                                    dispatch.status = "sent"
                                    dispatch.sent_at = datetime.utcnow()
                                    sent_count += 1
                                    status = "sent"
                                    _persist_campaign_message(db_session, phone, message, campaign.name, channel_id=_base_channel_id, campaign_id=campaign.id)
                                    break
                                else:
                                    error_code = result.get("error_code", "UNKNOWN")
                                    error_msg = result.get("error", "Erro desconhecido")
                                    print(f"[DISPATCH-FAIL] canal={_base_channel_id} assessor={assessor.get('email','')} motivo={error_code} detalhe={str(error_msg)[:200]} api_response={str(result)[:300]}")
                                    
                                    is_retryable = (
                                        error_code.startswith("HTTP_5") or 
                                        "500" in error_code or
                                        "502" in error_code or
                                        "503" in error_code or
                                        error_code in ["TIMEOUT", "CONNECTION_ERROR", "HTTP_ERROR"]
                                    )
                                    
                                    if is_retryable and attempt < MAX_RETRY_ATTEMPTS:
                                        await asyncio.sleep(RETRY_DELAY_SECONDS)
                                        attempt += 1
                                        continue
                                    else:
                                        dispatch.status = "failed"
                                        dispatch.error_message = error_msg
                                        dispatch.error_details = translate_error_to_natural_language(error_code, error_msg, phone)
                                        failed_count += 1
                                        status = "failed"
                                        break
                            except Exception as e:
                                error_msg = str(e)
                                if attempt < MAX_RETRY_ATTEMPTS:
                                    await asyncio.sleep(RETRY_DELAY_SECONDS)
                                    attempt += 1
                                else:
                                    dispatch.status = "failed"
                                    dispatch.error_message = error_msg
                                    dispatch.error_details = f"Erro de conexao: {error_msg}"
                                    failed_count += 1
                                    status = "failed"
                                    break
                    elif not _chan_inactive_base:
                        if not phone:
                            dispatch.status = "failed"
                            dispatch.error_message = "Telefone não informado"
                            dispatch.error_details = "O assessor não possui telefone WhatsApp cadastrado"
                            failed_count += 1
                            status = "failed"
                            error_msg = "Telefone não informado"
                        elif not _base_act_cfg:
                            dispatch.status = "simulated"
                            dispatch.error_details = "Disparo simulado - Z-API não configurado"
                            dispatch.sent_at = datetime.utcnow()
                            sent_count += 1
                            status = "simulated"
                    
                    db_session.commit()
                    
                    percent = round((current_index / total_assessors) * 100, 1)
                    progress_data = {
                        'type': 'progress',
                        'current': current_index,
                        'total': total_assessors,
                        'percent': percent,
                        'assessor_name': assessor_name,
                        'assessor_phone': phone,
                        'status': status,
                        'error': error_msg,
                        'sent_count': sent_count,
                        'failed_count': failed_count,
                        'attempts_made': attempt
                    }
                    yield f"data: {json.dumps(progress_data, ensure_ascii=False)}\n\n"
                    
                finally:
                    db_session.close()
                
                if current_index < total_assessors:
                    delay = get_random_dispatch_delay()
                    await asyncio.sleep(delay)
        
        except asyncio.CancelledError:
            cancelled = True
        finally:
            if campaign.id in cancelled_campaigns:
                del cancelled_campaigns[campaign.id]
            
            db_final = SessionLocal()
            try:
                campaign_final = db_final.query(Campaign).filter(Campaign.id == campaign.id).first()
                if campaign_final:
                    if cancelled or campaign_final.status == "cancelling":
                        campaign_final.status = "cancelled"
                    elif campaign_final.status in [CampaignStatus.PROCESSING.value, "processing", "sending"]:
                        campaign_final.status = CampaignStatus.SENT.value
                    campaign_final.messages_sent = sent_count
                    campaign_final.messages_failed = failed_count
                    campaign_final.sent_at = datetime.utcnow()
                    db_final.commit()
            finally:
                db_final.close()
        
        if not cancelled:
            complete_data = {
                'type': 'complete',
                'total': total_assessors,
                'sent_count': sent_count,
                'failed_count': failed_count
            }
            yield f"data: {json.dumps(complete_data)}\n\n"
    
    return StreamingResponse(
        generate_events(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"
        }
    )


@router.post("/{campaign_id}/cancel")
async def cancel_campaign_dispatch(
    campaign_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    """Cancela o disparo de uma campanha em andamento."""
    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campanha não encontrada")
    
    if campaign.status not in [CampaignStatus.PROCESSING.value, "processing", "sending"]:
        raise HTTPException(status_code=409, detail="A campanha não está em andamento")
    
    cancelled_campaigns[campaign_id] = True
    
    campaign.status = "cancelling"
    db.commit()
    
    return {"success": True, "message": "Solicitação de cancelamento enviada"}


@router.get("/{campaign_id}/delivery-report")
async def get_campaign_delivery_report(
    campaign_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    """Task #302 — Diagnóstico de entrega de campanha.

    Para cada dispatch com status 'sent', verifica se existe WhatsAppMessage
    e Conversation correspondentes no banco, e se o channel_id bate.
    Útil para identificar onde mensagens de campanha estão se perdendo.
    """
    from database.models import WhatsAppMessage, Conversation
    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campanha não encontrada")

    dispatches = (
        db.query(CampaignDispatch)
        .filter(CampaignDispatch.campaign_id == campaign_id)
        .order_by(CampaignDispatch.id.asc())
        .all()
    )

    items = []
    for d in dispatches:
        clean_phone = ''.join(filter(str.isdigit, d.assessor_phone or ""))

        msg = None
        conv = None
        if clean_phone:
            msg = (
                db.query(WhatsAppMessage)
                .filter(
                    WhatsAppMessage.campaign_id == campaign_id,
                    WhatsAppMessage.phone == clean_phone,
                )
                .first()
            )
            # Task #302 — lookup por (phone, channel_id) quando canal conhecido, para evitar
            # falso "mismatch" em cenários multicanal onde o mesmo telefone existe em canais distintos.
            conv_query = db.query(Conversation).filter(Conversation.phone == clean_phone)
            if d.channel_id:
                conv = conv_query.filter(Conversation.channel_id == d.channel_id).first()
                if conv is None:
                    # fallback: qualquer conversa com esse telefone (canal ainda não setado)
                    conv = conv_query.first()
            else:
                conv = conv_query.first()

        channel_match = None
        if msg and conv and d.channel_id is not None:
            channel_match = (msg.channel_id == d.channel_id and conv.channel_id == d.channel_id)

        items.append({
            "dispatch_id": d.id,
            "phone": d.assessor_phone,
            "status": d.status,
            "channel_id": d.channel_id,
            "sent_at": d.sent_at.isoformat() if d.sent_at else None,
            "error_message": d.error_message,
            "has_whatsapp_message": msg is not None,
            "message_id": msg.id if msg else None,
            "message_channel_id": msg.channel_id if msg else None,
            "has_conversation": conv is not None,
            "conversation_id": conv.id if conv else None,
            "conversation_channel_id": conv.channel_id if conv else None,
            "conversation_last_message_at": conv.last_message_at.isoformat() if conv and conv.last_message_at else None,
            "channel_id_match": channel_match,
        })

    sent = [i for i in items if i["status"] == "sent"]
    failed = [i for i in items if i["status"] == "failed"]
    pending = [i for i in items if i["status"] not in ("sent", "failed")]
    missing_msg = [i for i in sent if not i["has_whatsapp_message"]]
    missing_conv = [i for i in sent if not i["has_conversation"]]
    channel_mismatch = [i for i in sent if i["channel_id_match"] is False]

    return {
        "campaign_id": campaign_id,
        "campaign_name": campaign.name,
        "campaign_status": campaign.status,
        "total_dispatches": len(items),
        "sent": len(sent),
        "failed": len(failed),
        "pending": len(pending),
        "issues": {
            "missing_whatsapp_message": len(missing_msg),
            "missing_conversation": len(missing_conv),
            "channel_id_mismatch": len(channel_mismatch),
        },
        # dispatches agrupados por status para facilitar análise
        "dispatches_sent": [i for i in items if i["status"] == "sent"],
        "dispatches_failed": failed,
        "dispatches_pending": pending,
    }


@router.get("/")
async def list_campaigns(
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    """Lista todas as campanhas (unificadas + legadas) com paginação."""
    from database.models import CadenceCampaign, CadenceCampaignContact
    from sqlalchemy import func as sql_func

    campaigns = db.query(Campaign).order_by(Campaign.created_at.desc()).offset(skip).limit(limit).all()

    result = []
    # Task #221 — para campanhas em cadência, pré-calcular próximo envio
    # de uma vez só (uma query por campanha) para mostrar no card.
    cadence_ids = [c.id for c in campaigns if (c.delivery_mode == "cadence"
                   or c.status in ("firing_cadence", "paused_cadence", "cadence_done"))]
    next_etas: dict = {}
    if cadence_ids:
        from database.models import CampaignDispatch
        rows = (
            db.query(CampaignDispatch.campaign_id, sql_func.min(CampaignDispatch.scheduled_for))
            .filter(
                CampaignDispatch.campaign_id.in_(cadence_ids),
                CampaignDispatch.status == "pending",
                CampaignDispatch.scheduled_for.isnot(None),
            )
            .group_by(CampaignDispatch.campaign_id)
            .all()
        )
        next_etas = {cid: dt for cid, dt in rows}

    for c in campaigns:
        result.append({
            "id": c.id,
            "name": c.name,
            "status": c.status,
            "original_filename": c.original_filename,
            "total_assessors": c.total_assessors,
            "total_recommendations": c.total_recommendations,
            "messages_sent": c.messages_sent,
            "messages_failed": c.messages_failed,
            "delivery_mode": c.delivery_mode or "immediate",
            "daily_limit": c.daily_limit,
            "deadline_days": c.deadline_days,
            "created_at": c.created_at.isoformat() if c.created_at else None,
            "sent_at": c.sent_at.isoformat() if c.sent_at else None,
            "template_name": c.template.name if c.template else None,
            "source": "unified",
            "next_send_eta": next_etas.get(c.id).isoformat() if next_etas.get(c.id) else None,
        })

    legacy_campaigns = db.query(CadenceCampaign).order_by(CadenceCampaign.created_at.desc()).limit(50).all()
    # Task #221 — próximo envio das legadas em uma query agregada
    legacy_next_etas: dict = {}
    if legacy_campaigns:
        from database.models import CadenceCampaignContact as _LCC
        legacy_ids = [lc.id for lc in legacy_campaigns]
        for cid, dt in (
            db.query(_LCC.campaign_id, sql_func.min(_LCC.scheduled_for))
            .filter(_LCC.campaign_id.in_(legacy_ids), _LCC.status == "pending",
                    _LCC.scheduled_for.isnot(None))
            .group_by(_LCC.campaign_id).all()
        ):
            legacy_next_etas[cid] = dt
    for lc in legacy_campaigns:
        sent_count = db.query(sql_func.count(CadenceCampaignContact.id)).filter(
            CadenceCampaignContact.campaign_id == lc.id,
            CadenceCampaignContact.status.in_(["sent", "responded"]),
        ).scalar() or 0
        failed_count = db.query(sql_func.count(CadenceCampaignContact.id)).filter(
            CadenceCampaignContact.campaign_id == lc.id,
            CadenceCampaignContact.status == "failed",
        ).scalar() or 0

        status_map = {"firing": "firing_cadence", "done": "cadence_done", "paused": "paused_cadence"}
        mapped_status = status_map.get(lc.status, lc.status)

        result.append({
            "id": -lc.id,
            "name": f"{lc.name} (legado)",
            "status": mapped_status,
            "original_filename": None,
            "total_assessors": lc.total_contacts or 0,
            "total_recommendations": 0,
            "messages_sent": sent_count,
            "messages_failed": failed_count,
            "delivery_mode": "cadence",
            "daily_limit": lc.daily_limit,
            "deadline_days": None,
            "created_at": lc.created_at.isoformat() if lc.created_at else None,
            "sent_at": lc.start_date.isoformat() if lc.start_date else None,
            "template_name": None,
            "source": "legacy_cadence",
            "next_send_eta": legacy_next_etas.get(lc.id).isoformat() if legacy_next_etas.get(lc.id) else None,
        })

    result.sort(key=lambda x: x.get("created_at") or "", reverse=True)
    return result


class CampaignStructureCreate(BaseModel):
    name: str
    ticker: Optional[str] = None
    structure_type: str
    campaign_slug: str
    key_data: Optional[dict] = None
    diagram_filename: Optional[str] = None
    material_id: Optional[int] = None
    valid_from: Optional[str] = None
    valid_until: Optional[str] = None


class CampaignStructureUpdate(BaseModel):
    name: Optional[str] = None
    ticker: Optional[str] = None
    structure_type: Optional[str] = None
    campaign_slug: Optional[str] = None
    key_data: Optional[dict] = None
    diagram_filename: Optional[str] = None
    material_id: Optional[int] = None
    valid_from: Optional[str] = None
    valid_until: Optional[str] = None
    is_active: Optional[bool] = None


def _structure_to_dict(s: CampaignStructure) -> dict:
    import json as _json
    return {
        "id": s.id,
        "name": s.name,
        "ticker": s.ticker,
        "structure_type": s.structure_type,
        "campaign_slug": s.campaign_slug,
        "key_data": _json.loads(s.key_data) if s.key_data else {},
        "diagram_filename": s.diagram_filename,
        "material_id": s.material_id,
        "valid_from": s.valid_from.isoformat() if s.valid_from else None,
        "valid_until": s.valid_until.isoformat() if s.valid_until else None,
        "is_active": s.is_active == 1,
        "created_at": s.created_at.isoformat() if s.created_at else None,
        "updated_at": s.updated_at.isoformat() if s.updated_at else None,
    }


@router.get("/structures")
async def list_campaign_structures(
    active_only: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    query = db.query(CampaignStructure)
    if active_only:
        now = datetime.utcnow()
        query = query.filter(
            CampaignStructure.is_active == 1,
            (CampaignStructure.valid_from.is_(None)) | (CampaignStructure.valid_from <= now),
            (CampaignStructure.valid_until.is_(None)) | (CampaignStructure.valid_until >= now),
        )
    structures = query.order_by(CampaignStructure.created_at.desc()).all()
    return {"structures": [_structure_to_dict(s) for s in structures]}


@router.get("/active-structures")
async def get_active_campaign_structures(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    now = datetime.utcnow()
    structures = db.query(CampaignStructure).filter(
        CampaignStructure.is_active == 1,
        (CampaignStructure.valid_from.is_(None)) | (CampaignStructure.valid_from <= now),
        (CampaignStructure.valid_until.is_(None)) | (CampaignStructure.valid_until >= now),
    ).order_by(CampaignStructure.name).all()
    return {"structures": [_structure_to_dict(s) for s in structures]}


@router.get("/derivative-slugs")
async def list_derivative_slugs(
    current_user: User = Depends(require_admin_or_gestao())
):
    from scripts.xpi_derivatives.derivatives_dataset import get_all_structures
    structures = get_all_structures()
    return {
        "slugs": [
            {"slug": s["slug"], "name": s["name"], "tab": s["tab"]}
            for s in structures
        ]
    }


@router.post("/structures")
async def create_campaign_structure(
    data: CampaignStructureCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    existing = db.query(CampaignStructure).filter(
        CampaignStructure.campaign_slug == data.campaign_slug
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Já existe uma estrutura com esse slug")

    structure = CampaignStructure(
        name=data.name,
        ticker=data.ticker,
        structure_type=data.structure_type,
        campaign_slug=data.campaign_slug,
        key_data=json.dumps(data.key_data) if data.key_data else "{}",
        diagram_filename=data.diagram_filename,
        material_id=data.material_id,
        valid_from=datetime.fromisoformat(data.valid_from) if data.valid_from else None,
        valid_until=datetime.fromisoformat(data.valid_until) if data.valid_until else None,
        is_active=1,
        created_by=int(current_user.id),
    )
    db.add(structure)
    db.commit()
    db.refresh(structure)
    return {"success": True, "structure": _structure_to_dict(structure)}


@router.get("/structures/{structure_id}")
async def get_campaign_structure(
    structure_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    structure = db.query(CampaignStructure).filter(CampaignStructure.id == structure_id).first()
    if not structure:
        raise HTTPException(status_code=404, detail="Estrutura não encontrada")
    return _structure_to_dict(structure)


@router.put("/structures/{structure_id}")
async def update_campaign_structure(
    structure_id: int,
    data: CampaignStructureUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    structure = db.query(CampaignStructure).filter(CampaignStructure.id == structure_id).first()
    if not structure:
        raise HTTPException(status_code=404, detail="Estrutura não encontrada")

    if data.name is not None:
        structure.name = data.name
    if data.ticker is not None:
        structure.ticker = data.ticker
    if data.structure_type is not None:
        structure.structure_type = data.structure_type
    if data.campaign_slug is not None:
        existing = db.query(CampaignStructure).filter(
            CampaignStructure.campaign_slug == data.campaign_slug,
            CampaignStructure.id != structure_id
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail="Slug já em uso")
        structure.campaign_slug = data.campaign_slug
    if data.key_data is not None:
        structure.key_data = json.dumps(data.key_data)
    if data.diagram_filename is not None:
        structure.diagram_filename = data.diagram_filename
    if data.material_id is not None:
        structure.material_id = data.material_id
    if data.valid_from is not None:
        structure.valid_from = datetime.fromisoformat(data.valid_from) if data.valid_from else None
    if data.valid_until is not None:
        structure.valid_until = datetime.fromisoformat(data.valid_until) if data.valid_until else None
    if data.is_active is not None:
        structure.is_active = 1 if data.is_active else 0

    db.commit()
    db.refresh(structure)
    return {"success": True, "structure": _structure_to_dict(structure)}


@router.delete("/structures/{structure_id}")
async def delete_campaign_structure(
    structure_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    structure = db.query(CampaignStructure).filter(CampaignStructure.id == structure_id).first()
    if not structure:
        raise HTTPException(status_code=404, detail="Estrutura não encontrada")

    structure.is_active = 0
    db.commit()
    return {"success": True, "message": "Estrutura desativada"}


@router.post("/structures/{structure_id}/diagram")
async def upload_campaign_diagram(
    structure_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    structure = db.query(CampaignStructure).filter(CampaignStructure.id == structure_id).first()
    if not structure:
        raise HTTPException(status_code=404, detail="Estrutura não encontrada")

    if not file.filename:
        raise HTTPException(status_code=400, detail="Nome do arquivo inválido")

    import os
    slug = structure.campaign_slug
    ext = os.path.splitext(file.filename)[1].lower() or ".png"
    diagram_name = f"{slug}{ext}"
    diagram_dir = os.path.join("static", "derivatives_diagrams")
    os.makedirs(diagram_dir, exist_ok=True)
    diagram_path = os.path.join(diagram_dir, diagram_name)

    contents = await file.read()
    with open(diagram_path, "wb") as f:
        f.write(contents)

    structure.diagram_filename = diagram_name
    db.commit()

    return {
        "success": True,
        "diagram_filename": diagram_name,
        "message": f"Diagrama salvo em {diagram_path}"
    }


@cadence_router.get("/recent-turbo-aborts")
@router.get("/cadence/recent-turbo-aborts")
async def get_recent_turbo_aborts(
    since_id: int = 0,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao()),
):
    """
    Task #222 — Lista eventos `turbo_aborted_safety` recentes (últimos
    10 min) com id > since_id, para o frontend exibir um toast vermelho
    com a razão e o perfil restaurado quando o freio do turbo dispara.
    Retorna até 20 eventos ordenados por id ascendente.
    """
    from database.models import CadenceCampaignEvent
    from datetime import timedelta as _td, timezone as _tz

    cutoff = datetime.now(_tz.utc) - _td(minutes=10)
    rows = (
        db.query(CadenceCampaignEvent)
        .filter(CadenceCampaignEvent.event_type == "turbo_aborted_safety")
        .filter(CadenceCampaignEvent.id > int(since_id or 0))
        .filter(CadenceCampaignEvent.occurred_at >= cutoff)
        .order_by(CadenceCampaignEvent.id.asc())
        .limit(20)
        .all()
    )
    out = []
    for r in rows:
        payload = r.payload or {}
        out.append({
            "id": int(r.id),
            "campaign_kind": r.campaign_kind,
            "campaign_id": int(r.campaign_id or 0),
            "occurred_at": r.occurred_at.isoformat() if r.occurred_at else None,
            "reason": payload.get("reason") or "—",
            "restored_profile": payload.get("restored_profile") or "conservador",
            "is_backfill": bool(getattr(r, "is_backfill", False)),
        })
    return {"events": out}


@cadence_router.get("/engine-state")
@router.get("/cadence/engine-state")
async def get_cadence_engine_state(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    """
    Estado normalizado do motor de cadência (task #221).

    DECLARADA ANTES de `/{campaign_id}` para evitar route-shadowing —
    `campaign_id: int` rejeitaria 'cadence' com 422 sem chegar aqui.

    Retorna `state` como enum normalizado:
      - `ok`                       — motor ativo, pronto para enviar.
      - `pause_anti_block`         — pausa por falhas consecutivas.
      - `out_of_business_hours`    — fora da janela 09-18 seg-sex.
      - `lunch_break`              — pausa de almoço 12:00-13:00.
      - `global_cooldown`          — esperando cooldown do último envio.

    Inclui `projected_resume_at` (datetime ISO) com a melhor estimativa
    do próximo momento em que o motor estará apto a enviar.
    """
    from services.cadence_events import (
        get_engine_state,
        ENGINE_STATE_OK,
        ENGINE_STATE_ANTI_BLOCK,
        ENGINE_STATE_OUT_OF_HOURS,
        ENGINE_STATE_LUNCH_BREAK,
        ENGINE_STATE_GLOBAL_COOLDOWN,
    )
    from services.cadence_profiles import get_profile
    from datetime import datetime as _dt, timedelta as _td
    from zoneinfo import ZoneInfo

    state = get_engine_state(db)
    tz = ZoneInfo("America/Sao_Paulo")
    now_local = _dt.now(tz)

    # Janelas
    is_weekend = now_local.weekday() >= 5
    is_before_work = now_local.hour < 9
    is_after_work = now_local.hour >= 18
    is_lunch = now_local.hour == 12
    in_business_hours = not (is_weekend or is_before_work or is_after_work)

    # Pausa anti-block persistente
    pause_until = state.get("pause_until")
    is_paused_anti_block = bool(pause_until) and pause_until > _dt.now(pause_until.tzinfo)

    # Cooldown global — usa o cooldown do perfil "conservador" como
    # estimativa segura (perfil mais lento). É só uma projeção visual;
    # o motor avalia o cooldown real por campanha em cada tick.
    last_send = state.get("last_send_at")
    cooldown_default = int(get_profile("conservador").get("cooldown_seconds", 480))
    in_global_cooldown = False
    cooldown_until = None
    if last_send:
        cd_end = last_send + _td(seconds=cooldown_default)
        if cd_end > _dt.now(last_send.tzinfo):
            in_global_cooldown = True
            cooldown_until = cd_end

    # Estado normalizado (precedência: pausa > horário > almoço > cooldown > ok)
    if is_paused_anti_block:
        normalized_state = ENGINE_STATE_ANTI_BLOCK
        normalized_reason = state.get("pause_reason") or "anti_block"
        projected_resume_at = pause_until
    elif not in_business_hours:
        normalized_state = ENGINE_STATE_OUT_OF_HOURS
        normalized_reason = "weekend" if is_weekend else ("before_work" if is_before_work else "after_work")
        projected_resume_at = _next_business_window_start(now_local)
    elif is_lunch:
        # Estado observado (12-13h BRT). O motor NÃO pausa de fato neste
        # intervalo (comportamento original preservado). projected_resume_at
        # é exposto para alimentar o contador do banner por observabilidade,
        # mas a UI rotula como informativo.
        normalized_state = ENGINE_STATE_LUNCH_BREAK
        normalized_reason = "informational"
        projected_resume_at = now_local.replace(hour=13, minute=0, second=0, microsecond=0)
    elif in_global_cooldown:
        normalized_state = ENGINE_STATE_GLOBAL_COOLDOWN
        normalized_reason = "cooldown"
        projected_resume_at = cooldown_until
    else:
        normalized_state = ENGINE_STATE_OK
        normalized_reason = None
        projected_resume_at = None

    return {
        "state": normalized_state,
        "reason": normalized_reason,
        "projected_resume_at": projected_resume_at.isoformat() if projected_resume_at else None,
        "last_tick_at": state.get("last_tick_at").isoformat() if state.get("last_tick_at") else None,
        "last_send_at": last_send.isoformat() if last_send else None,
        "pause_until": pause_until.isoformat() if pause_until else None,
        "pause_reason": state.get("pause_reason"),
        # Compat com banner V1 (mantido):
        "is_paused_now": is_paused_anti_block,
        "consecutive_failures": int(state.get("consecutive_failures") or 0),
        "in_business_hours_now": in_business_hours,
        "in_lunch_break_now": is_lunch,
        "in_global_cooldown_now": in_global_cooldown,
        "cooldown_seconds_default": cooldown_default,
        "now": now_local.isoformat(),
        "business_hours": {"start": "09:00", "end": "18:00", "weekdays_only": True, "lunch_break": "12:00-13:00"},
    }


def _next_business_window_start(now_local):
    """Próximo horário em que o motor estará dentro da janela comercial.
    Considera fim de semana e antes/depois do expediente.
    """
    from datetime import timedelta as _td
    candidate = now_local
    # Hoje antes das 9h e dia útil → hoje 9h
    if candidate.weekday() < 5 and candidate.hour < 9:
        return candidate.replace(hour=9, minute=0, second=0, microsecond=0)
    # Senão, avança 1 dia até cair em dia útil e retorna 9h
    candidate = (candidate + _td(days=1)).replace(hour=9, minute=0, second=0, microsecond=0)
    while candidate.weekday() >= 5:
        candidate = candidate + _td(days=1)
    return candidate


@router.get("/{campaign_id}")
async def get_campaign(
    campaign_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    """Busca uma campanha por ID."""
    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campanha não encontrada")
    
    dispatches = db.query(CampaignDispatch).filter(
        CampaignDispatch.campaign_id == campaign_id
    ).all()
    
    try:
        column_mapping = json.loads(str(campaign.column_mapping)) if campaign.column_mapping else {}
        custom_fields_mapping = json.loads(str(campaign.custom_fields_mapping)) if campaign.custom_fields_mapping else {}
        processed_data = json.loads(str(campaign.processed_data)) if campaign.processed_data else []
    except json.JSONDecodeError:
        column_mapping = {}
        custom_fields_mapping = {}
        processed_data = []
    
    file_columns = []
    if processed_data and len(processed_data) > 0:
        file_columns = list(processed_data[0].keys())
    
    return {
        "id": campaign.id,
        "name": campaign.name,
        "status": campaign.status,
        "template_id": campaign.template_id,
        "template_name": campaign.template.name if campaign.template else None,
        "custom_template_content": campaign.custom_template_content,
        "original_filename": campaign.original_filename,
        "column_mapping": column_mapping,
        "custom_fields_mapping": custom_fields_mapping,
        "file_columns": file_columns,
        "total_assessors": campaign.total_assessors,
        "total_recommendations": campaign.total_recommendations,
        "messages_sent": campaign.messages_sent,
        "messages_failed": campaign.messages_failed,
        "created_at": campaign.created_at.isoformat() if campaign.created_at else None,
        "sent_at": campaign.sent_at.isoformat() if campaign.sent_at else None,
        "has_data": len(processed_data) > 0,
        "dispatches": [
            {
                "assessor_id": d.assessor_id,
                "assessor_name": d.assessor_name,
                "assessor_phone": d.assessor_phone,
                "status": d.status,
                "error_message": d.error_message,
                "sent_at": d.sent_at.isoformat() if d.sent_at else None,
                "message_content": d.message_content if d.message_content else None
            }
            for d in dispatches
        ]
    }


@router.get("/{campaign_id}/failures")
async def get_campaign_failures(
    campaign_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    """
    Retorna análise detalhada das falhas de uma campanha.
    Agrupa falhas por tipo e fornece descrição em linguagem natural.
    """
    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campanha não encontrada")
    
    failed_dispatches = db.query(CampaignDispatch).filter(
        CampaignDispatch.campaign_id == campaign_id,
        CampaignDispatch.status == "failed"
    ).all()
    
    if not failed_dispatches:
        return {
            "campaign_id": campaign_id,
            "campaign_name": campaign.name,
            "total_failures": 0,
            "failures": [],
            "summary": "Nenhuma falha registrada nesta campanha."
        }
    
    failures = []
    error_categories = {}
    
    for d in failed_dispatches:
        error_msg = d.error_message or "Erro desconhecido"
        error_detail = d.error_details or translate_error_to_natural_language("UNKNOWN", error_msg, d.assessor_phone or "")
        
        category = "Outro"
        if "timeout" in error_msg.lower():
            category = "Timeout"
        elif "connection" in error_msg.lower() or "conectar" in error_msg.lower():
            category = "Conexao"
        elif "401" in error_msg or "403" in error_msg or "credenciais" in error_detail.lower():
            category = "Autenticacao"
        elif "telefone" in error_msg.lower() or "phone" in error_msg.lower() or "numero" in error_msg.lower():
            category = "Numero Invalido"
        elif "session" in error_msg.lower() or "sessao" in error_detail.lower():
            category = "Sessao WhatsApp"
        
        if category not in error_categories:
            error_categories[category] = 0
        error_categories[category] += 1
        
        api_response_parsed = None
        if d.api_response:
            try:
                api_response_parsed = json.loads(d.api_response)
            except json.JSONDecodeError:
                api_response_parsed = d.api_response
        
        failures.append({
            "assessor_name": d.assessor_name or "Desconhecido",
            "assessor_phone": d.assessor_phone or "Nao informado",
            "error_message": error_msg,
            "error_details": error_detail,
            "category": category,
            "api_response": api_response_parsed
        })
    
    summary_parts = []
    for cat, count in sorted(error_categories.items(), key=lambda x: -x[1]):
        if cat == "Conexao":
            summary_parts.append(f"{count} falha(s) de conexao com o servidor Z-API")
        elif cat == "Autenticacao":
            summary_parts.append(f"{count} falha(s) de autenticacao (chave de API)")
        elif cat == "Numero Invalido":
            summary_parts.append(f"{count} numero(s) de telefone invalido(s) ou ausente(s)")
        elif cat == "Sessao WhatsApp":
            summary_parts.append(f"{count} problema(s) com a sessao do WhatsApp")
        elif cat == "Timeout":
            summary_parts.append(f"{count} timeout(s) - servidor demorou para responder")
        else:
            summary_parts.append(f"{count} outro(s) erro(s)")
    
    summary = "Resumo das falhas: " + "; ".join(summary_parts) + "." if summary_parts else "Nenhuma falha categorizada."
    
    return {
        "campaign_id": campaign_id,
        "campaign_name": campaign.name,
        "total_failures": len(failed_dispatches),
        "categories": error_categories,
        "summary": summary,
        "failures": failures
    }


@router.delete("/{campaign_id}")
async def delete_campaign(
    campaign_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    """Remove uma campanha e seus dispatches."""
    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campanha não encontrada")
    
    if campaign.status == CampaignStatus.PROCESSING.value:
        raise HTTPException(status_code=400, detail="Não é possível excluir uma campanha em processamento")
    
    db.delete(campaign)
    db.commit()
    
    return {"message": "Campanha excluída com sucesso"}


@router.get("/{campaign_id}/debug")
async def debug_campaign(
    campaign_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    """
    Endpoint de diagnóstico para verificar dados da campanha.
    """
    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campanha não encontrada")
    
    template_content = DEFAULT_TEMPLATE_CONTENT
    template_source = "default"
    
    if campaign.custom_template_content:
        template_content = str(campaign.custom_template_content)
        template_source = "custom"
    elif campaign.template_id:
        template = db.query(MessageTemplate).filter(MessageTemplate.id == campaign.template_id).first()
        if template:
            template_content = str(template.content)
            template_source = f"template_{campaign.template_id}"
    
    try:
        column_mapping = json.loads(str(campaign.column_mapping)) if campaign.column_mapping else {}
        custom_mapping = json.loads(str(campaign.custom_fields_mapping)) if campaign.custom_fields_mapping else {}
        data = json.loads(str(campaign.processed_data)) if campaign.processed_data else []
    except json.JSONDecodeError as e:
        return {"error": f"JSON decode error: {str(e)}"}
    
    grouped = {}
    if column_mapping and data:
        grouped = group_recommendations_by_assessor(data, column_mapping, custom_mapping, db)
    
    content_line_template = campaign.message_content_template or ""
    sample_message = ""
    if grouped:
        first_key = list(grouped.keys())[0]
        sample_message = build_message(template_content, grouped[first_key], custom_mapping, content_line_template, group_by_client=bool(campaign.group_by_client))
    
    return {
        "campaign_id": campaign_id,
        "campaign_name": campaign.name,
        "status": campaign.status,
        "template_source": template_source,
        "template_content_preview": template_content[:500] if template_content else None,
        "template_has_nome_assessor": "{{nome_assessor}}" in template_content if template_content else False,
        "template_has_lista_clientes": "{{lista_clientes}}" in template_content if template_content else False,
        "column_mapping": column_mapping,
        "custom_mapping": custom_mapping,
        "data_rows_count": len(data),
        "data_first_row": data[0] if data else None,
        "data_first_row_keys": list(data[0].keys()) if data else [],
        "grouped_assessors_count": len(grouped),
        "grouped_keys": list(grouped.keys())[:5],
        "first_assessor_data": grouped[list(grouped.keys())[0]] if grouped else None,
        "sample_message_preview": sample_message[:500] if sample_message else None
    }


def _persist_campaign_message(
    db_session: Session,
    phone: str,
    message: str,
    campaign_name: str = "",
    channel_id: Optional[int] = None,
    campaign_id: Optional[int] = None,
):
    """Persiste conversa e mensagem de campanha após envio bem-sucedido.

    Task #287 — propaga channel_id, assessor_id e campaign_id para garantir
    roteamento correto das respostas do agente (resolve_channel_client_for_conversation).
    Task #302 — seta last_message_at (para ordenação na Central) e emite SSE.
    """
    if not phone or not message:
        return
    _conv_id_for_sse = None
    try:
        from database.models import WhatsAppMessage, MessageDirection, MessageType, SenderType, Conversation, Assessor as _Assessor
        clean_phone = ''.join(filter(str.isdigit, phone))
        if not clean_phone:
            return

        now = datetime.utcnow()

        conversation = db_session.query(Conversation).filter(
            Conversation.phone == clean_phone
        ).first()

        _created_new = False
        if not conversation:
            conversation = Conversation(
                phone=clean_phone,
                channel_id=channel_id,
                ticket_status="new",
                last_message_at=now,
            )
            db_session.add(conversation)
            db_session.flush()
            _created_new = True
        else:
            if channel_id and not conversation.channel_id:
                conversation.channel_id = channel_id
            if not conversation.ticket_status:
                conversation.ticket_status = "new"
            # Task #302 — garantir que last_message_at reflita o envio da campanha
            conversation.last_message_at = now
            db_session.flush()

        if _created_new and not conversation.assessor_id:
            try:
                _assessor = (
                    db_session.query(_Assessor)
                    .filter(_Assessor.telefone_whatsapp == clean_phone)
                    .first()
                )
                if _assessor:
                    conversation.assessor_id = _assessor.id
                    db_session.flush()
            except Exception as _lookup_err:
                print(f"[CAMPAIGN_MSG] Aviso: lookup de assessor falhou para telefone={clean_phone} canal={channel_id}: {_lookup_err}")

        # Task #310 — Reset de estado de conversa ao disparar campanha.
        # Garante que a resposta do assessor não seja interceptada por estado
        # residual de sessões anteriores (stalled_interactions ou awaiting_confirmation).
        # Proteção: skip quando há atendimento humano ativo (ticket_status = 'open').
        #
        # Cobre conversas novas (defaults já corretos) e existentes. Para conversas
        # novas os defaults (stalled=0, awaiting=False) garantem que o bloco é no-op.
        if conversation.ticket_status != "open":
            from database.models import ConversationState as _CS
            _prev_stalled = conversation.stalled_interactions or 0
            _prev_awaiting = bool(conversation.awaiting_confirmation)
            _any_reset = False

            _reset_parts = []
            if _prev_stalled >= 3:
                conversation.stalled_interactions = 0
                _any_reset = True
                _reset_parts.append(f"stalled={_prev_stalled}→0")
            if _prev_awaiting:
                conversation.awaiting_confirmation = False
                conversation.confirmation_sent_at = None
                _any_reset = True
                _reset_parts.append("awaiting_confirmation=True→False")
            if _any_reset:
                conversation.conversation_state = _CS.READY.value
                print(
                    f"[CAMPAIGN] Conversa {conversation.id} resetada: "
                    + ", ".join(_reset_parts)
                )

        # Task #311 — Corrige channel_id da conversa para o canal da campanha.
        # Garante que respostas do assessor sejam roteadas pelo canal correto.
        # Guards: (a) ticket_status = 'open' — atendimento humano ativo, não interferir;
        #         (b) assessor tem channel_id direto configurado — canal do assessor prevalece.
        if channel_id and conversation.channel_id != channel_id and conversation.ticket_status != "open":
            # Fail-safe: se não for possível confirmar ausência de override do assessor,
            # assume que há override e preserva conversation.channel_id (canal do assessor
            # é a fonte de verdade). Só atualiza quando a verificação é conclusiva (sem exceção).
            _assessor_has_explicit_channel = False   # False = "não tem override" (padrão seguro)
            _lookup_failed = False

            if conversation.assessor_id:
                try:
                    _ov_channel = (
                        db_session.query(_Assessor.channel_id)
                        .filter(_Assessor.id == conversation.assessor_id)
                        .scalar()
                    )
                    if _ov_channel:
                        _assessor_has_explicit_channel = True
                except Exception as _lookup_exc:
                    # Não podemos confirmar a ausência de override — comportamento conservador:
                    # preserva conversation.channel_id para não rotear pelo canal errado.
                    _lookup_failed = True
                    print(
                        f"[CAMPAIGN] Aviso: lookup de override de canal para assessor_id="
                        f"{conversation.assessor_id} falhou — channel_id não atualizado "
                        f"(fail-safe): {type(_lookup_exc).__name__}: {_lookup_exc}"
                    )

            if not _assessor_has_explicit_channel and not _lookup_failed:
                _old_channel_id = conversation.channel_id
                conversation.channel_id = channel_id
                print(
                    f"[CAMPAIGN] Conversa {conversation.id} channel_id atualizado: "
                    f"{_old_channel_id} → {channel_id}"
                )

        tag = f"[Campanha: {campaign_name}] " if campaign_name else ""
        record = WhatsAppMessage(
            chat_id=clean_phone,
            phone=clean_phone,
            direction=MessageDirection.OUTBOUND.value,
            message_type=MessageType.TEXT.value,
            from_me=True,
            body=f"{tag}{message}",
            ai_response=None,
            ai_intent="campaign_dispatch",
            sender_type=SenderType.BOT.value,
            conversation_id=conversation.id,
            channel_id=channel_id,
            is_from_campaign=True,
            campaign_id=campaign_id,
        )
        db_session.add(record)
        db_session.commit()
        _conv_id_for_sse = conversation.id
        print(f"[CAMPAIGN_MSG] Mensagem salva: telefone={clean_phone} canal={channel_id} campanha_id={campaign_id} conv_id={_conv_id_for_sse}")
    except Exception as e:
        print(f"[CAMPAIGN_MSG] Erro ao salvar mensagem de campanha: {e}")
        try:
            db_session.rollback()
        except Exception:
            pass

    # Task #302 — emite notificação SSE para a Central de Conversas atualizar em tempo real.
    # Executado fora do try/except principal para não afetar o commit já realizado.
    if _conv_id_for_sse:
        try:
            import asyncio
            from services.sse_manager import get_sse_manager
            _sse = get_sse_manager()
            asyncio.get_event_loop().create_task(
                _sse.notify_new_message(_conv_id_for_sse, {
                    "phone": phone,
                    "channel_id": channel_id,
                    "campaign_id": campaign_id,
                    "direction": "outbound",
                    "sender_type": "bot",
                    "is_from_campaign": True,
                    "last_message_at": datetime.utcnow().isoformat(),
                })
            )
        except Exception as _sse_err:
            print(f"[CAMPAIGN_MSG] Aviso: falha ao emitir SSE para conv_id={_conv_id_for_sse}: {_sse_err}")


class CadenceDispatchRequest(BaseModel):
    # Optional: quando None, o limite efetivo é derivado do perfil
    # (50 conservador / 80 padrão / 120 acelerado). Mantém compatibilidade
    # com chamadas que continuam enviando 50 explicitamente.
    daily_limit: Optional[int] = None
    deadline_days: int = 5
    cadence_profile: str = "conservador"

    @validator('daily_limit')
    def validate_daily_limit(cls, v):
        if v is None:
            return v
        if v < 1:
            raise ValueError('Limite diário deve ser pelo menos 1')
        if v > 500:
            raise ValueError('Limite diário não pode exceder 500')
        return v

    @validator('deadline_days')
    def validate_deadline_days(cls, v):
        if v < 1:
            raise ValueError('Prazo deve ser pelo menos 1 dia')
        if v > 60:
            raise ValueError('Prazo não pode exceder 60 dias')
        return v

    @validator('cadence_profile')
    def validate_cadence_profile(cls, v):
        from services.cadence_profiles import is_valid_profile, PROFILES, USER_SELECTABLE_PROFILES
        if not is_valid_profile(v):
            raise ValueError(f'Perfil inválido. Opções: {", ".join(USER_SELECTABLE_PROFILES)}')
        return str(v).strip().lower()


class CadenceProfileChangeRequest(BaseModel):
    cadence_profile: str

    @validator('cadence_profile')
    def validate_cadence_profile(cls, v):
        from services.cadence_profiles import is_valid_profile, PROFILES, USER_SELECTABLE_PROFILES
        if not is_valid_profile(v):
            raise ValueError(f'Perfil inválido. Opções: {", ".join(USER_SELECTABLE_PROFILES)}')
        return str(v).strip().lower()


class CadenceFinalizeNowRequest(BaseModel):
    """Task #222 — payload do modo "Finalizar disparos agora" (turbo seguro).

    ``confirmation`` deve ser exatamente "FINALIZAR" para evitar acionamento
    acidental. ``override_business_hours`` é opt-in: quando True, o
    cronograma turbo não respeita 09-18h ao distribuir os horários (mesmo
    assim, o motor de envio mantém a janela comercial — o override sinaliza
    que sobras serão enviadas no próximo dia útil).
    """
    confirmation: str
    override_business_hours: bool = False

    @validator('confirmation')
    def validate_confirmation(cls, v):
        if (v or "").strip().upper() != "FINALIZAR":
            raise ValueError('Confirmação inválida — digite exatamente "FINALIZAR".')
        return "FINALIZAR"


@router.post("/{campaign_id}/dispatch-cadence")
async def dispatch_campaign_cadence(
    campaign_id: int,
    data: CadenceDispatchRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    from services.campaign_planner import calculate_daily_plan, _get_business_days, _build_daily_schedule
    from zoneinfo import ZoneInfo
    import math

    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campanha não encontrada")

    if campaign.status == CampaignStatus.SENT.value:
        raise HTTPException(status_code=400, detail="Esta campanha já foi enviada")

    # Task #222 — quando a campanha está em turbo, o reagendamento via
    # dispatch-cadence sobrescreveria o perfil sem limpar a flag, criando
    # estado contraditório (turbo ativo + perfil normal). Bloqueia até o
    # operador finalizar/cancelar o turbo.
    if bool(getattr(campaign, "cadence_turbo_active", False)):
        raise HTTPException(
            status_code=409,
            detail="Campanha em modo turbo — finalize os disparos atuais antes de reagendar.",
        )

    existing_dispatches = db.query(CampaignDispatch).filter(
        CampaignDispatch.campaign_id == campaign_id,
        CampaignDispatch.status.in_(["pending", "processing"]),
    ).count()
    if existing_dispatches > 0:
        raise HTTPException(
            status_code=409,
            detail=f"Esta campanha já possui {existing_dispatches} disparos pendentes. Aguarde a conclusão ou cancele antes de reagendar."
        )

    source_type = getattr(campaign, 'source_type', 'upload') or 'upload'

    header_template = campaign.message_header or ""
    content_template = campaign.message_content_template or ""
    footer_template = campaign.message_footer or ""
    use_blocks = bool(header_template.strip() or content_template.strip() or footer_template.strip())

    template_content = DEFAULT_TEMPLATE_CONTENT
    if not use_blocks:
        if campaign.custom_template_content:
            candidate = str(campaign.custom_template_content)
            if template_has_required_variables(candidate):
                template_content = candidate
        elif campaign.template_id:
            template = db.query(MessageTemplate).filter(MessageTemplate.id == campaign.template_id).first()
            if template:
                candidate = str(template.content)
                if template_has_required_variables(candidate):
                    template_content = candidate

    try:
        column_mapping = json.loads(str(campaign.column_mapping)) if campaign.column_mapping else {}
        custom_mapping = json.loads(str(campaign.custom_fields_mapping)) if campaign.custom_fields_mapping else {}
        processed_data = json.loads(str(campaign.processed_data)) if campaign.processed_data else []
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Erro nos dados da campanha")

    if source_type in ["base", "base_assessores"]:
        assessor_list = processed_data
        if not assessor_list:
            raise HTTPException(status_code=400, detail="Nenhum assessor encontrado nos dados da campanha")

        base_header = header_template
        base_content = content_template
        base_footer = footer_template

        if not base_header and not base_content and not base_footer:
            if campaign.custom_template_content:
                base_content = str(campaign.custom_template_content)
            elif campaign.template_id:
                tmpl = db.query(MessageTemplate).filter(MessageTemplate.id == campaign.template_id).first()
                if tmpl:
                    base_content = str(tmpl.content)
            else:
                base_content = "Ola, {{nome_assessor}}!"

        dispatches_data = []

        # Task #224 — resolver channel_id por unidade para assessores da base.
        _base_unidades = {assessor.get("unidade", "") for assessor in assessor_list if assessor.get("unidade")}
        _base_channel_map_by_unidade: dict = {}
        if _base_unidades:
            from database.models import UnidadeChannelMapping as _UCM_BASE
            for _ucm_b in db.query(_UCM_BASE).filter(_UCM_BASE.unidade.in_(_base_unidades)).all():
                _base_channel_map_by_unidade[_ucm_b.unidade] = _ucm_b.channel_id

        def _resolve_ch_from_assessor_dict(a: dict) -> int | None:
            """Retorna channel_id para um assessor da base (Task #224)."""
            _direct = a.get("channel_id") or a.get("zapi_channel_id")
            if _direct:
                return int(_direct)
            _unid = a.get("unidade", "")
            return _base_channel_map_by_unidade.get(_unid)

        for assessor in assessor_list:
            assessor_name = assessor.get("nome", "")
            phone = assessor.get("telefone_whatsapp", "") or assessor.get("telefone", "")
            variables = build_assessor_variables(assessor)

            message_parts = []
            header_rendered = replace_variables_generic(base_header, variables)
            if header_rendered.strip():
                message_parts.append(header_rendered.strip())
            content_rendered = replace_variables_generic(base_content, variables)
            if content_rendered.strip():
                message_parts.append(content_rendered.strip())
            footer_rendered = replace_variables_generic(base_footer, variables)
            if footer_rendered.strip():
                message_parts.append(footer_rendered.strip())
            message = "\n\n".join(message_parts)
            message = re.sub(r'\{\{[^}]+\}\}', '', message)

            dispatches_data.append({
                "assessor_id": str(assessor.get("id", "")),
                "assessor_email": assessor.get("email", ""),
                "assessor_phone": phone,
                "assessor_name": assessor_name,
                "message_content": message,
                "channel_id": _resolve_ch_from_assessor_dict(assessor),
            })
    else:
        grouped = group_recommendations_by_assessor(processed_data, column_mapping, custom_mapping, db)
        if not grouped:
            raise HTTPException(status_code=400, detail="Nenhum assessor encontrado para disparo")

        dispatches_data = []
        content_line_template = campaign.message_content_template or ""
        is_grouped = bool(campaign.group_by_client)

        # Task #224 — channel map para assessores do path de upload.
        _upload_emails = [ad.get("email_assessor", "") for ad in grouped.values() if ad.get("email_assessor")]
        _upload_channel_map = _batch_resolve_channels(_upload_emails, db) if _upload_emails else {}

        for assessor_id, assessor_data in grouped.items():
            if use_blocks:
                wrapper_parts = []
                if header_template.strip():
                    wrapper_parts.append(header_template.strip())
                if content_template.strip() and "{{lista_clientes}}" in content_template:
                    wrapper_parts.append(content_template.strip())
                else:
                    wrapper_parts.append("{{lista_clientes}}")
                if footer_template.strip():
                    wrapper_parts.append(footer_template.strip())
                wrapper_template = "\n\n".join(wrapper_parts)
                message = build_message(wrapper_template, assessor_data, custom_mapping, content_line_template, group_by_client=is_grouped)
            else:
                message = build_message(template_content, assessor_data, custom_mapping, content_line_template, group_by_client=is_grouped)

            phone = assessor_data.get("telefone", "")
            dispatches_data.append({
                "assessor_id": assessor_id,
                "assessor_email": assessor_data.get("email_assessor", ""),
                "assessor_phone": phone,
                "assessor_name": assessor_data.get("nome_assessor", ""),
                "message_content": message,
                "channel_id": _upload_channel_map.get(assessor_data.get("email_assessor", "")),
            })

    if not dispatches_data:
        raise HTTPException(status_code=400, detail="Nenhum destinatário encontrado")

    # Task #312 — Pre-flight check: verifica conectividade e webhook antes de criar dispatches.
    _pf_cad_cids = list({d.get("channel_id") for d in dispatches_data})
    _pf_cad_result = await _run_preflight_check(_pf_cad_cids, db)
    if not _pf_cad_result["all_ok"]:
        _pf_cad_bad = [ch for ch in _pf_cad_result["channels"] if not ch.get("ok")]
        raise HTTPException(
            status_code=422,
            detail={
                "preflight_failed": True,
                "message": "Um ou mais canais de envio apresentam problemas. Corrija antes de agendar.",
                "channels": _pf_cad_bad,
            },
        )

    # Resolve daily_limit efetivo: se o usuário não informou, usa o valor
    # do perfil de cadência selecionado (50/80/120).
    from services.cadence_profiles import get_profile as _get_profile
    profile_cfg_for_limit = _get_profile(data.cadence_profile)
    daily_limit = data.daily_limit if data.daily_limit is not None else int(profile_cfg_for_limit["daily_limit"])
    deadline_days = data.deadline_days

    tz = ZoneInfo("America/Sao_Paulo")
    now = datetime.now(tz)
    today = now.date()

    from database.models import Conversation, WhatsAppMessage
    from datetime import timedelta

    for d in dispatches_data:
        phone = d["assessor_phone"]
        if phone:
            phone_suffix = phone[-8:] if len(phone) >= 8 else phone
            recent_msg = (
                db.query(WhatsAppMessage.id)
                .join(Conversation, Conversation.id == WhatsAppMessage.conversation_id)
                .filter(
                    Conversation.phone.ilike(f"%{phone_suffix}%"),
                    WhatsAppMessage.direction == "INBOUND",
                    WhatsAppMessage.created_at >= now - timedelta(days=2),
                )
                .first()
            )
            if recent_msg:
                d["priority"] = 1
                continue
            any_conv = db.query(Conversation.id).filter(
                Conversation.phone.ilike(f"%{phone_suffix}%")
            ).first()
            if any_conv:
                d["priority"] = 2
            else:
                d["priority"] = 3
        else:
            d["priority"] = 3

    dispatches_data = [d for d in dispatches_data if d.get("message_content", "").strip()]
    if not dispatches_data:
        raise HTTPException(status_code=400, detail="Nenhum dispatch com conteudo valido. Verifique o template da campanha.")

    dispatches_data.sort(key=lambda x: x.get("priority", 3))

    total = len(dispatches_data)
    plan = calculate_daily_plan(total, deadline_days, daily_limit)
    business_days = _get_business_days(today, deadline_days)
    daily_cap = min(daily_limit, math.ceil(total / len(business_days))) if business_days else daily_limit

    p3_daily_limit = 15
    dispatch_idx = 0

    for day in business_days:
        if dispatch_idx >= total:
            break
        day_batch = []
        p3_count = 0
        for i in range(dispatch_idx, total):
            if len(day_batch) >= daily_cap:
                break
            d = dispatches_data[i]
            if d.get("priority") == 3:
                if p3_count >= p3_daily_limit:
                    continue
                p3_count += 1
            day_batch.append(d)

        times = _build_daily_schedule(len(day_batch), day, profile=data.cadence_profile)
        fallback_time = datetime.combine(day, datetime.min.time().replace(hour=9), tzinfo=tz)

        for j, d in enumerate(day_batch):
            sched_time = times[j] if j < len(times) else fallback_time
            dispatch = CampaignDispatch(
                campaign_id=campaign_id,
                assessor_id=d["assessor_id"],
                assessor_email=d.get("assessor_email", ""),
                assessor_phone=d["assessor_phone"],
                assessor_name=d.get("assessor_name", ""),
                message_content=d["message_content"],
                status="pending",
                scheduled_for=sched_time,
                priority=d.get("priority", 3),
                channel_id=d.get("channel_id"),
            )
            db.add(dispatch)
            dispatch_idx += 1

    overflow = dispatches_data[dispatch_idx:]
    if overflow:
        extra_day = business_days[-1] + timedelta(days=1) if business_days else today + timedelta(days=1)
        extra_days = _get_business_days(extra_day, math.ceil(len(overflow) / daily_cap) + 1)
        ov_idx = 0
        for ed in extra_days:
            if ov_idx >= len(overflow):
                break
            batch = overflow[ov_idx:ov_idx + daily_cap]
            times = _build_daily_schedule(len(batch), ed, profile=data.cadence_profile)
            fallback_time = datetime.combine(ed, datetime.min.time().replace(hour=9), tzinfo=tz)
            for j, d in enumerate(batch):
                sched_time = times[j] if j < len(times) else fallback_time
                dispatch = CampaignDispatch(
                    campaign_id=campaign_id,
                    assessor_id=d["assessor_id"],
                    assessor_email=d.get("assessor_email", ""),
                    assessor_phone=d["assessor_phone"],
                    assessor_name=d.get("assessor_name", ""),
                    message_content=d["message_content"],
                    status="pending",
                    scheduled_for=sched_time,
                    priority=d.get("priority", 3),
                    channel_id=d.get("channel_id"),
                )
                db.add(dispatch)
            ov_idx += len(batch)

    campaign.delivery_mode = "cadence"
    # Persistimos exatamente o que o usuário enviou (None preserva fallback
    # dinâmico do perfil em mudanças futuras de cadence_profile).
    campaign.daily_limit = data.daily_limit
    campaign.deadline_days = deadline_days
    campaign.cadence_profile = data.cadence_profile
    campaign.status = "firing_cadence"
    campaign.total_assessors = total
    campaign.sent_at = now
    db.commit()

    # Task #221 — eventos de criação e início (cadência unificada)
    from services.cadence_events import (
        emit_event as _obs_emit, CAMPAIGN_KIND_UNIFIED,
        EVENT_CAMPAIGN_CREATED as _EVT_CREATED,
        EVENT_CAMPAIGN_STARTED as _EVT_STARTED,
    )
    _user_id_actor = getattr(current_user, "id", None)
    _obs_emit(db, CAMPAIGN_KIND_UNIFIED, campaign.id, _EVT_CREATED, {
        "name": campaign.name,
        "total_contacts": int(total),
        "daily_limit": daily_limit,
        "deadline_days": deadline_days,
        "cadence_profile": campaign.cadence_profile,
    }, user_id=_user_id_actor)
    _obs_emit(db, CAMPAIGN_KIND_UNIFIED, campaign.id, _EVT_STARTED, {
        "auto": True,
    }, user_id=_user_id_actor)

    print(f"[CADENCE] Campanha '{campaign.name}' (id={campaign_id}) agendada com cadência: {total} contatos, {deadline_days} dias, limite {daily_limit}/dia")

    return {
        "message": "Campanha agendada com cadência",
        "total_contacts": total,
        "daily_limit": daily_limit,
        "deadline_days": deadline_days,
        "plano": plan,
        "alerta": plan.get("alerta"),
    }


@router.get("/{campaign_id}/channel-preview")
async def get_campaign_channel_preview(
    campaign_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao()),
):
    """Task #224 Fix 5 — prévia de canais para Step 3 (pré-disparo).
    Lê processed_data e resolve channel_id por email de assessor via
    _batch_resolve_channels, retornando a mesma estrutura de /channel-summary
    para que a UI exiba os cards antes do primeiro envio.
    """
    from database.models import ZAPIChannel
    from sqlalchemy import func as sql_func

    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campanha não encontrada")

    try:
        proc = json.loads(str(campaign.processed_data)) if campaign.processed_data else []
    except Exception:
        proc = []

    # Extrai emails: suporta campos 'email_assessor' (upload/grouped) e 'email' (base).
    emails: list[str] = []
    for row in proc:
        e = (
            row.get("email_assessor")
            or row.get("email")
            or row.get("Email")
            or row.get("assessor_email")
            or ""
        )
        if e:
            emails.append(str(e).strip().lower())

    emails = [e for e in emails if e]

    channel_map = _batch_resolve_channels(list(set(emails)), db) if emails else {}

    # Agrupa: channel_id → {emails únicos, contagem total}.
    channel_groups: dict = {}
    for email in emails:
        ch_id = channel_map.get(email)
        if ch_id not in channel_groups:
            channel_groups[ch_id] = {"emails": set(), "total": 0}
        channel_groups[ch_id]["emails"].add(email)
        channel_groups[ch_id]["total"] += 1

    # Se não há dados de processed_data ainda, retorna lista vazia.
    if not channel_groups:
        return {"campaign_id": campaign_id, "channels": [], "source": "preview"}

    # Resolve unidades dos assessores agrupados por canal.
    all_unique_emails = list({e for g in channel_groups.values() for e in g["emails"]})
    assessor_unidade_map: dict[str, str | None] = {}
    if all_unique_emails:
        for a in db.query(Assessor.email, Assessor.unidade).filter(
            Assessor.email.in_(all_unique_emails)
        ).all():
            assessor_unidade_map[a.email] = a.unidade

    # Enriquece com labels dos canais em batch.
    channel_ids = [cid for cid in channel_groups if cid is not None]
    channels_by_id: dict = {}
    if channel_ids:
        for ch in db.query(ZAPIChannel).filter(ZAPIChannel.id.in_(channel_ids)).all():
            channels_by_id[ch.id] = ch

    summary = []
    for ch_id, grp in channel_groups.items():
        ch = channels_by_id.get(ch_id) if ch_id else None
        unidades = sorted({
            assessor_unidade_map.get(e)
            for e in grp["emails"]
            if assessor_unidade_map.get(e)
        })
        summary.append({
            "channel_id": ch_id,
            "channel_label": ch.label if ch else "Canal legado (env vars)",
            "channel_phone": ch.phone_number if ch else None,
            "channel_is_active": ch.is_active if ch else True,
            "total": grp["total"],
            "assessor_count": len(grp["emails"]),
            "unidades": unidades,
            "sent": 0,
            "failed": 0,
            "pending": grp["total"],
        })

    # Ordena pelo total decrescente para exibição consistente.
    summary.sort(key=lambda x: x["total"], reverse=True)

    return {"campaign_id": campaign_id, "channels": summary, "source": "preview"}


@router.get("/{campaign_id}/channel-summary")
async def get_campaign_channel_summary(
    campaign_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao()),
):
    """Task #224 — resume de canais usados em uma campanha.
    Agrega os dispatches por channel_id e retorna label + phone_number
    do canal Z-API correspondente.
    """
    from database.models import ZAPIChannel
    from sqlalchemy import func as sql_func

    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campanha não encontrada")

    from sqlalchemy import case as sa_case

    rows = (
        db.query(
            CampaignDispatch.channel_id,
            sql_func.count(CampaignDispatch.id).label("total"),
            sql_func.sum(
                sa_case((CampaignDispatch.status == "sent", 1), else_=0)
            ).label("sent"),
            sql_func.sum(
                sa_case((CampaignDispatch.status == "failed", 1), else_=0)
            ).label("failed"),
            sql_func.sum(
                sa_case((CampaignDispatch.status == "pending", 1), else_=0)
            ).label("pending"),
        )
        .filter(CampaignDispatch.campaign_id == campaign_id)
        .group_by(CampaignDispatch.channel_id)
        .all()
    )

    # Busca labels dos canais em batch.
    channel_ids = [r.channel_id for r in rows if r.channel_id is not None]
    channels_by_id = {}
    if channel_ids:
        for ch in db.query(ZAPIChannel).filter(ZAPIChannel.id.in_(channel_ids)).all():
            channels_by_id[ch.id] = ch

    # Busca assessor_count e unidades por canal via dispatches (Task #224 Fix B).
    # Cada dispatch tem assessor_phone; usamos email do assessor para agrupar.
    ch_id_to_phones: dict = {}
    for d in (
        db.query(CampaignDispatch.channel_id, CampaignDispatch.assessor_phone)
        .filter(
            CampaignDispatch.campaign_id == campaign_id,
            CampaignDispatch.assessor_phone.isnot(None),
        )
        .distinct()
        .all()
    ):
        ch_id_to_phones.setdefault(d.channel_id, set()).add(d.assessor_phone)

    # Resolve unidades via Assessor.telefone_whatsapp (últimos 10 dígitos).
    all_phones = {p for phones in ch_id_to_phones.values() for p in phones}
    phone_to_unidade: dict[str, str | None] = {}
    if all_phones:
        def _norm10_cs(p: str) -> str:
            return p.lstrip("+").replace(" ", "").replace("-", "")[-10:]
        from sqlalchemy import or_ as _or_cs
        _cs_filters = [Assessor.telefone_whatsapp.ilike(f"%{_norm10_cs(p)}%") for p in all_phones]
        for a in db.query(Assessor.telefone_whatsapp, Assessor.unidade).filter(
            _or_cs(*_cs_filters)
        ).all():
            phone_to_unidade[_norm10_cs(a.telefone_whatsapp or "")] = a.unidade

    summary = []
    for r in rows:
        ch = channels_by_id.get(r.channel_id) if r.channel_id else None
        phones_for_ch = ch_id_to_phones.get(r.channel_id, set())
        unidades = sorted({
            phone_to_unidade.get(_norm10_cs(p) if all_phones else p)
            for p in phones_for_ch
            if phone_to_unidade.get(_norm10_cs(p) if all_phones else p)
        }) if phones_for_ch else []
        summary.append({
            "channel_id": r.channel_id,
            "channel_label": ch.label if ch else "Canal legado (env vars)",
            "channel_phone": ch.phone_number if ch else None,
            "channel_is_active": ch.is_active if ch else True,
            "total": int(r.total or 0),
            "assessor_count": len(phones_for_ch),
            "unidades": unidades,
            "sent": int(r.sent or 0),
            "failed": int(r.failed or 0),
            "pending": int(r.pending or 0),
        })

    return {"campaign_id": campaign_id, "channels": summary}


@router.get("/{campaign_id}/cadence-status")
async def get_cadence_status(
    campaign_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    from sqlalchemy import func as sql_func

    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campanha não encontrada")

    sent = db.query(sql_func.count(CampaignDispatch.id)).filter(
        CampaignDispatch.campaign_id == campaign_id,
        CampaignDispatch.status == "sent"
    ).scalar() or 0

    pending = db.query(sql_func.count(CampaignDispatch.id)).filter(
        CampaignDispatch.campaign_id == campaign_id,
        CampaignDispatch.status == "pending"
    ).scalar() or 0

    failed = db.query(sql_func.count(CampaignDispatch.id)).filter(
        CampaignDispatch.campaign_id == campaign_id,
        CampaignDispatch.status == "failed"
    ).scalar() or 0

    responded = db.query(sql_func.count(CampaignDispatch.id)).filter(
        CampaignDispatch.campaign_id == campaign_id,
        CampaignDispatch.status == "responded"
    ).scalar() or 0

    total_delivered = sent + responded
    response_rate = round((responded / total_delivered * 100), 1) if total_delivered > 0 else 0.0

    from zoneinfo import ZoneInfo
    tz = ZoneInfo("America/Sao_Paulo")

    daily_stats = (
        db.query(
            sql_func.date(CampaignDispatch.sent_at).label("day"),
            sql_func.count(CampaignDispatch.id).label("sent"),
        )
        .filter(
            CampaignDispatch.campaign_id == campaign_id,
            CampaignDispatch.status.in_(["sent", "responded"]),
            CampaignDispatch.sent_at.isnot(None),
        )
        .group_by(sql_func.date(CampaignDispatch.sent_at))
        .order_by(sql_func.date(CampaignDispatch.sent_at).asc())
        .all()
    )

    responded_by_day = (
        db.query(
            sql_func.date(CampaignDispatch.responded_at).label("day"),
            sql_func.count(CampaignDispatch.id).label("responded"),
        )
        .filter(
            CampaignDispatch.campaign_id == campaign_id,
            CampaignDispatch.status == "responded",
            CampaignDispatch.responded_at.isnot(None),
        )
        .group_by(sql_func.date(CampaignDispatch.responded_at))
        .order_by(sql_func.date(CampaignDispatch.responded_at).asc())
        .all()
    )

    responded_map = {str(r.day): r.responded for r in responded_by_day}

    daily_log = []
    for row in daily_stats:
        day_str = str(row.day)
        daily_log.append({
            "date": day_str,
            "sent": row.sent,
            "responded": responded_map.get(day_str, 0),
        })

    from services.cadence_profiles import list_profiles, get_profile

    # Task #221 — KPI "última hora" e "próximo envio" + último erro Z-API.
    from datetime import datetime as _dt, timedelta as _td
    now_utc = _dt.utcnow()
    one_hour_ago = now_utc - _td(hours=1)
    sent_last_hour = db.query(sql_func.count(CampaignDispatch.id)).filter(
        CampaignDispatch.campaign_id == campaign_id,
        CampaignDispatch.status.in_(["sent", "responded"]),
        CampaignDispatch.sent_at >= one_hour_ago,
    ).scalar() or 0

    next_dispatch = (
        db.query(CampaignDispatch)
        .filter(
            CampaignDispatch.campaign_id == campaign_id,
            CampaignDispatch.status == "pending",
        )
        .order_by(CampaignDispatch.scheduled_for.asc())
        .first()
    )
    next_send_eta = next_dispatch.scheduled_for.isoformat() if (next_dispatch and next_dispatch.scheduled_for) else None

    last_err_row = (
        db.query(CampaignDispatch)
        .filter(
            CampaignDispatch.campaign_id == campaign_id,
            CampaignDispatch.last_error_message.isnot(None),
        )
        .order_by(CampaignDispatch.id.desc())
        .first()
    )
    last_error_message = last_err_row.last_error_message if last_err_row else None

    profile_cfg = get_profile(getattr(campaign, "cadence_profile", None))

    return {
        "id": campaign.id,
        "name": campaign.name,
        "status": campaign.status,
        "delivery_mode": campaign.delivery_mode or "immediate",
        "total_contacts": campaign.total_assessors or 0,
        "sent": sent + responded,
        "pending": pending,
        "failed": failed,
        "responded": responded,
        "response_rate": response_rate,
        "sent_last_hour": int(sent_last_hour),
        "next_send_eta": next_send_eta,
        "last_error_message": last_error_message,
        "cooldown_seconds": int(profile_cfg.get("cooldown_seconds", 0)),
        "daily_limit": campaign.daily_limit,
        "deadline_days": campaign.deadline_days,
        "cadence_profile": getattr(campaign, "cadence_profile", None) or "conservador",
        # Task #222 — flags do modo "Finalizar disparos agora" para a UI
        "cadence_turbo_active": bool(getattr(campaign, "cadence_turbo_active", False)),
        "cadence_turbo_origin_profile": getattr(campaign, "cadence_turbo_origin_profile", None),
        "available_profiles": list_profiles(),
        "daily_log": daily_log,
    }


# ============================================================================
# Task #221 — Endpoints de observabilidade do motor de cadência
# ============================================================================

@router.get("/{campaign_id}/events")
async def get_campaign_events(
    campaign_id: int,
    limit: int = 100,
    before_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    """Lista eventos de timeline de uma campanha unificada (Task #221)."""
    from services.cadence_events import list_events_for_campaign, CAMPAIGN_KIND_UNIFIED
    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campanha não encontrada")
    safe_limit = max(1, min(int(limit), 500))
    events = list(list_events_for_campaign(db, CAMPAIGN_KIND_UNIFIED, campaign_id, limit=safe_limit, before_id=before_id))
    return {
        "campaign_id": campaign_id,
        "campaign_kind": CAMPAIGN_KIND_UNIFIED,
        "count": len(events),
        "events": events,
    }


@router.patch("/{campaign_id}/cadence-profile")
async def change_cadence_profile(
    campaign_id: int,
    data: CadenceProfileChangeRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    """
    Troca o perfil de velocidade da cadência de uma campanha unificada e
    reagenda automaticamente apenas os dispatches com status 'pending'.
    Não altera os já enviados, em processamento ou falhos.
    """
    from services.campaign_planner import reschedule_unified_pending_dispatches

    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campanha não encontrada")

    # Task #222 — bloqueia troca manual enquanto a campanha está em modo turbo
    # para evitar estado inconsistente (flag turbo + perfil normal selecionado).
    if bool(getattr(campaign, "cadence_turbo_active", False)):
        raise HTTPException(
            status_code=409,
            detail="Campanha em modo turbo — não é possível trocar o perfil agora. Aguarde a conclusão dos disparos.",
        )

    old_profile = getattr(campaign, "cadence_profile", None) or "conservador"
    campaign.cadence_profile = data.cadence_profile
    db.commit()

    rescheduled = 0
    if campaign.status in ("firing_cadence", "paused_cadence"):
        rescheduled = reschedule_unified_pending_dispatches(campaign_id, db)

    from services.cadence_events import emit_event as _obs_emit, CAMPAIGN_KIND_UNIFIED, EVENT_PROFILE_CHANGED as _EVT
    _obs_emit(db, CAMPAIGN_KIND_UNIFIED, campaign.id, _EVT, {
        "old_profile": old_profile,
        "new_profile": data.cadence_profile,
        "rescheduled_count": int(rescheduled),
    }, user_id=getattr(current_user, "id", None))

    print(
        f"[CADENCE] Perfil da campanha '{campaign.name}' (id={campaign_id}) alterado: "
        f"{old_profile} → {data.cadence_profile} ({rescheduled} dispatches reagendados)"
    )

    return {
        "message": "Perfil atualizado",
        "cadence_profile": data.cadence_profile,
        "rescheduled_dispatches": rescheduled,
    }


@router.patch("/{campaign_id}/cadence-pause")
async def pause_cadence_campaign(
    campaign_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campanha não encontrada")
    if campaign.status != "firing_cadence":
        raise HTTPException(status_code=400, detail=f"Campanha não pode ser pausada (status: {campaign.status})")

    campaign.status = "paused_cadence"
    db.commit()
    from services.cadence_events import emit_event as _obs_emit, CAMPAIGN_KIND_UNIFIED, EVENT_CAMPAIGN_PAUSED as _EVT
    _obs_emit(db, CAMPAIGN_KIND_UNIFIED, campaign.id, _EVT, {"manual": True}, user_id=getattr(current_user, "id", None))
    print(f"[CADENCE] Campanha '{campaign.name}' (id={campaign_id}) pausada")
    return {"message": "Campanha pausada", "status": "paused_cadence"}


@router.patch("/{campaign_id}/cadence-resume")
async def resume_cadence_campaign(
    campaign_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    from services.campaign_planner import _get_business_days, _build_daily_schedule
    from services.cadence_profiles import get_profile
    from zoneinfo import ZoneInfo
    from datetime import timedelta
    import math

    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campanha não encontrada")
    if campaign.status != "paused_cadence":
        raise HTTPException(status_code=400, detail=f"Campanha não pode ser retomada (status: {campaign.status})")

    pending_dispatches = (
        db.query(CampaignDispatch)
        .filter(
            CampaignDispatch.campaign_id == campaign_id,
            CampaignDispatch.status == "pending",
        )
        .order_by(CampaignDispatch.priority.asc())
        .all()
    )

    if pending_dispatches:
        # Task #222 — campanhas em modo turbo retomam usando o cronograma turbo
        # (30-90s) em vez do scheduler padrão de minutos.
        if bool(getattr(campaign, "cadence_turbo_active", False)):
            from services.campaign_planner import reschedule_unified_for_turbo
            reschedule_unified_for_turbo(
                campaign.id,
                db,
                override_business_hours=False,
            )
        else:
            # Usa o perfil atual da campanha — assim retomadas após troca de perfil
            # respeitam a velocidade configurada e o limite diário sugerido.
            profile_name = getattr(campaign, "cadence_profile", None) or "conservador"
            profile_cfg = get_profile(profile_name)
            tz = ZoneInfo("America/Sao_Paulo")
            today = datetime.now(tz).date()
            daily_limit = campaign.daily_limit or int(profile_cfg["daily_limit"])
            deadline_days = campaign.deadline_days or 5
            business_days = _get_business_days(today, deadline_days)
            daily_cap = min(daily_limit, math.ceil(len(pending_dispatches) / len(business_days))) if business_days else daily_limit

            idx = 0
            for day in business_days:
                if idx >= len(pending_dispatches):
                    break
                batch = pending_dispatches[idx:idx + daily_cap]
                times = _build_daily_schedule(len(batch), day, profile=profile_name)
                for j, d in enumerate(batch):
                    if j < len(times):
                        d.scheduled_for = times[j]
                idx += len(batch)

    campaign.status = "firing_cadence"
    db.commit()
    from services.cadence_events import emit_event as _obs_emit, CAMPAIGN_KIND_UNIFIED, EVENT_CAMPAIGN_RESUMED as _EVT
    _obs_emit(db, CAMPAIGN_KIND_UNIFIED, campaign.id, _EVT, {"manual": True}, user_id=getattr(current_user, "id", None))
    print(f"[CADENCE] Campanha '{campaign.name}' (id={campaign_id}) retomada")
    return {"message": "Campanha retomada", "status": "firing_cadence"}


@router.post("/{campaign_id}/cadence-finalize-now")
async def finalize_unified_cadence_now(
    campaign_id: int,
    data: CadenceFinalizeNowRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_gestao())
):
    """
    Task #222 — Aciona o modo "Finalizar disparos agora" (turbo seguro)
    para uma campanha unificada em ``firing_cadence`` ou ``paused_cadence``.

    Comprime os dispatches pendentes com intervalo 30-90s, soft cap 150/dia,
    cooldown global 30s. Mantém defesas anti-bloqueio mínimas: janela
    comercial (com override opt-in), pausa 20min após 2 falhas Z-API
    consecutivas e freio automático que reverte para o perfil de origem
    em caso de falhas seguidas.

    Idempotente: se já estiver em turbo, apenas reagenda os pendentes.
    """
    from services.campaign_planner import reschedule_unified_for_turbo
    from services.cadence_profiles import TURBO_PROFILE_NAME
    from services.cadence_events import (
        emit_event as _obs_emit, CAMPAIGN_KIND_UNIFIED, EVENT_TURBO_STARTED,
    )

    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campanha não encontrada")
    if campaign.status not in ("firing_cadence", "paused_cadence"):
        raise HTTPException(
            status_code=400,
            detail=f"Apenas campanhas em disparo podem ser finalizadas (status: {campaign.status})",
        )

    pending_count = (
        db.query(CampaignDispatch)
        .filter(
            CampaignDispatch.campaign_id == campaign_id,
            CampaignDispatch.status == "pending",
        )
        .count()
    )
    if pending_count == 0:
        raise HTTPException(status_code=400, detail="Sem dispatches pendentes para finalizar.")

    origin_profile = campaign.cadence_turbo_origin_profile or (
        getattr(campaign, "cadence_profile", None) or "conservador"
    )
    campaign.cadence_turbo_origin_profile = origin_profile
    campaign.cadence_turbo_active = True
    # Task #222 — persiste o override para que o motor saiba bypassar a
    # janela 09-18h apenas para esta campanha entre ticks/restarts.
    campaign.cadence_turbo_override_business_hours = bool(data.override_business_hours)
    campaign.cadence_profile = TURBO_PROFILE_NAME
    if campaign.status == "paused_cadence":
        campaign.status = "firing_cadence"
    db.commit()

    rescheduled = reschedule_unified_for_turbo(
        campaign_id, db, override_business_hours=bool(data.override_business_hours)
    )

    # Task #222 — ETA aproximada: pending × intervalo médio (60s).
    eta_seconds = int(pending_count) * 60

    _obs_emit(db, CAMPAIGN_KIND_UNIFIED, campaign.id, EVENT_TURBO_STARTED, {
        "original_profile": origin_profile,
        "origin_profile": origin_profile,  # alias retrocompatível
        "override_business_hours": bool(data.override_business_hours),
        "rescheduled_count": int(rescheduled),
        "pending_at_start": int(pending_count),
        "pending_count": int(pending_count),
        "eta_seconds": int(eta_seconds),
    }, user_id=getattr(current_user, "id", None))

    print(
        f"[CADENCE-TURBO] Campanha unificada '{campaign.name}' (id={campaign_id}) "
        f"em modo turbo. {rescheduled} dispatches reagendados "
        f"(origin={origin_profile}, override_horario={data.override_business_hours})."
    )

    return {
        "message": "Modo turbo ativado",
        "status": campaign.status,
        "cadence_profile": campaign.cadence_profile,
        "original_profile": origin_profile,
        "origin_profile": origin_profile,
        "rescheduled_dispatches": rescheduled,
        "pending_count": int(pending_count),
        "eta_seconds": int(eta_seconds),
        "override_business_hours": bool(data.override_business_hours),
    }
